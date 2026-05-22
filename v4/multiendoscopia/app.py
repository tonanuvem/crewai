""" app-endoscopia-v7
Interface Streamlit para Sistema de Controle de Procedimentos de Endoscopia

Instalar:
    pip install streamlit crewai python-dotenv pypdf openpyxl

Executar:
    streamlit run app_endoscopia.py
"""

import os
import logging
import queue
import threading
import time
from datetime import datetime
from pathlib import Path

import io
import re

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from docx import Document
from pypdf import PdfReader

# CrewAI
from crewai import Agent, Task, Crew, Process, LLM

# =============================================================================
# CONFIGURAÇÃO DE VARIÁVEIS DE AMBIENTE E LOG
# =============================================================================

load_dotenv()

def _resolve_log_path() -> Path:
    candidates = [Path("logs"), Path("/tmp/endoscopia_logs")]
    for candidate in candidates:
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            test_file = candidate / ".write_test"
            test_file.touch()
            test_file.unlink()
            return candidate
        except OSError:
            continue
    return Path("/tmp")

log_dir = _resolve_log_path()

_log_handlers: list[logging.Handler] = [logging.StreamHandler()]
try:
    _log_handlers.insert(0, logging.FileHandler(log_dir / "crew_logs.log", encoding="utf-8"))
except OSError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=_log_handlers,
    force=True,
)
logger = logging.getLogger(__name__)


# =============================================================================
# VALORES PADRÃO DOS AGENTES E TASKS
# =============================================================================

ANALISTA_DEFAULTS = {
    "role": "Analista de Controle dos procedimentos de Endoscopia",
    "goal": """
- Retornar APENAS o CSV puro, sem markdown, sem explicações.
- Primeira linha obrigatoriamente o cabeçalho:
- Padronizar Data para DD/MM/AAAA.
- Padronizar nomes de paciente em MAIÚSCULAS sem espaços extras.
- Identificar o tipo pelo nome da planilha: PRODUCAO ou REPASSE
- Coluna TipoArquivo = "PRODUCAO" ou "REPASSE".

Desmembramento de Procedimentos Adicionais: 
- Se houver preenchimento na coluna "PROCEDIMENTOS ADICIONAIS" (ex: TESTE DE UREASE), você deve gerar uma LINHA EXTRA no CSV para cada procedimento listado.
- Na linha extra, o nome do procedimento adicional deve entrar na coluna Procedimento. As demais informações (Data, Atendimento, Paciente, Convênio, etc.) devem ser replicadas da linha principal.
- Para essas linhas extras, inclua obrigatoriamente a tag PROCEDIMENTO_ADICIONAL na coluna Observacao. A linha do procedimento principal deve ter a Observacao vazia.
""",
    "backstory": """
Expert em análise de faturamento hospitalar e auditoria de convênios médicos.
Especializado na Terminologia Unificada da Saúde Suplementar (TUSS) e nas regras de
faturamento dos principais convênios do Brasil.
Conhece em profundidade os campos das planilhas de produtividade de equipes de enfermagem
e os arquivos de repasse hospitalares (formato Hospital São Camilo e similares).
Sabe que o número de atendimento pode divergir entre os dois arquivos para o mesmo
paciente/procedimento, portanto não usa esse campo como chave única.
Seu principal objetivo é estruturar os dados com máxima fidelidade para que o agente
correlacionador consiga fazer o batimento correto.
    """,
    "task_description_template": """
Retorne APENAS o CSV puro, sem markdown, sem explicações.
Cabeçalho obrigatório na primeira linha
Não omita nenhuma linha — cada procedimento registrado deve constar no CSV gerado.
=== CONTEÚDO DO ARQUIVO ===
{conteudo_arquivo}
    """,
    "task_expected_output": (
        "CSV puro com cabeçalho na primeira linha, todas as linhas do arquivo estruturadas e coluna TipoArquivo preenchida com PRODUCAO ou REPASSE."
    ),
}

CORRELACIONADOR_DEFAULTS = {
    "role": "Correlacionador dos registros de endoscopia",
    "goal": """
Receber os CSVs padronizados gerados pelo Analista — um do tipo PRODUCAO e outro do tipo REPASSE — e realizar o Batimento Automático linha a linha, considerando as seguintes regras:

- A primeira linha deve ser obrigatoriamente o cabeçalho, indicando de qual planilha veio aquela coluna (acrescentar sufixo _PRODUCAO ou _REPASSE)
- NÃO remover nenhuma linha da PRODUCAO; Ordenar por Data crescente; Padronizar Data para DD/MM/AAAA.
- Padronizar ValorLiberado: ponto decimal, sem símbolo de moeda. 

- A saída do CSV_CORRELACAO deve conter todas as linhas do CSV de PRODUÇÃO como a base da sua resposta, onde devem ser incluídas as novas colunas com o relacionamento do CSV de REPASSE, sendo que para cada linha da PRODUCAO, deve tentar encontrar correspondência no REPASSE usando a chave composta com 3 campos (data + paciente + procedimento). Devem ser incluídas em cada linha somente as colunas do CSV de REPASSE que ainda não exista no CSV de PRODUÇÃO, e também deve ser incluída uma nova coluna de StatusCorrelacao, considerando a seguinte regra:
   a. Se ValorLiberado coincidir (ou REPASSE não tiver valor): StatusCorrelacao = "CORRELACIONADO"
   b. Se ValorLiberado for menor no REPASSE: StatusCorrelacao = "CORRELACIONADO_COM_DIVERGENCIA_VALOR"
   c. Se ValorLiberado = 0 no REPASSE: StatusCorrelacao = "CORRELACIONADO_COM_GLOSA_TOTAL"
   d. Se ValorLiberado > 0 mas menor que esperado: StatusCorrelacao = "CORRELACIONADO_COM_GLOSA_PARCIAL"
   e. Se NÃO encontrar correspondência no REPASSE: StatusCorrelacao = "NAO_FATURADO_NO_REPASSE"

- Para linhas do REPASSE sem correspondência na PRODUCAO, inserir no final do arquivo CSV_CORRELACAO as linhas como valores zerados do PRODUCAO e os valores existentes da PRODUÇÃO, e com nova coluna StatusCorrelacao = "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO"

- Para fazer esse relacionamento, deve-se usar a chave de correlação COMPOSTA pelos 3 campos a seguir:
  1. Data (DD/MM/AAAA) — tolerância de ± 1 dia para o mesmo episódio
  2. Paciente (nome normalizado em maiúsculas)
  3. Procedimento (correspondência semântica: "COLONOSCOPIA" bate com "Colonoscopia (Inclui A Retossigmoidoscopia)", "TESTE DE UREASE" bate com "Pesquisa de H. pylori - Teste da Urease", etc.)
    """,
    "backstory": """
Especialista em auditoria de faturamento médico, ETL e reconciliação de dados hospitalares.
Usa correspondência semântica para não perder correlações válidas.
Nunca inventa procedimentos — apenas padroniza, correlaciona e sinaliza divergências.
Seu trabalho alimenta diretamente a cobrança estruturada junto ao hospital e convênios.
    """,
    "task_description_template": """
Você recebeu os seguintes blocos CSV padronizados pelo Analista de Endoscopia, identificados como PRODUCAO e REPASSE. Execute o batimento automático conforme as regras do seu objetivo (goal), usando chave composta de Data + Paciente + Procedimento.
- Retorne APENAS o CSV correlacionado
- A primeira linha deve ser obrigatoriamente o cabeçalho
=== CONTEÚDO ===
{blocos}
    """,
    "task_expected_output": (
        "CSV puro com cabeçalho na primeira linha, todas as linhas da PRODUCAO como base, colunas complementares do REPASSE adicionadas e coluna StatusCorrelacao preenchida em cada linha."
    ),
}


def _init_agent_session_state():
    """
    Inicializa o session_state com os valores padrão dos agentes,
    caso ainda não existam (chamado uma vez no início do main).
    """
    if "analista_cfg" not in st.session_state:
        st.session_state["analista_cfg"] = dict(ANALISTA_DEFAULTS)
    if "correlacionador_cfg" not in st.session_state:
        st.session_state["correlacionador_cfg"] = dict(CORRELACIONADOR_DEFAULTS)


def _get_analista_cfg() -> dict:
    return st.session_state.get("analista_cfg", ANALISTA_DEFAULTS)


def _get_correlacionador_cfg() -> dict:
    return st.session_state.get("correlacionador_cfg", CORRELACIONADOR_DEFAULTS)


# =============================================================================
# HANDLER DE LOG EM TEMPO REAL PARA O STREAMLIT
# =============================================================================

class StreamlitLogHandler(logging.Handler):
    def __init__(self, log_queue: queue.Queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record: logging.LogRecord):
        try:
            self.log_queue.put_nowait(self.format(record))
        except queue.Full:
            pass


def render_log_line(line: str, container) -> None:
    line_lower = line.lower()
    if any(k in line_lower for k in ("error", "erro", "exception", "traceback")):
        container.error(f"🔴 {line}")
    elif any(k in line_lower for k in ("warning", "warn", "aviso")):
        container.warning(f"🟡 {line}")
    elif any(k in line_lower for k in ("task", "tarefa", "iniciando", "starting")):
        container.info(f"📋 {line}")
    elif any(k in line_lower for k in ("agent", "agente", "thinking", "pensando")):
        container.info(f"🤖 {line}")
    elif any(k in line_lower for k in ("action", "ação", "tool", "ferramenta")):
        container.info(f"⚙️ {line}")
    elif any(k in line_lower for k in ("final answer", "resposta final", "finished", "concluído", "complete")):
        container.success(f"✅ {line}")
    else:
        container.text(f"   {line}")


# =============================================================================
# FUNÇÕES DE LEITURA DE ARQUIVOS
# =============================================================================

def read_word_file(file) -> str:
    try:
        doc = Document(file)
        text = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(text)
    except Exception as e:
        st.error(f"Erro ao ler arquivo Word: {e}")
        return None


def read_pdf_file(file) -> str:
    try:
        pdf_reader = PdfReader(file)
        text = [
            page.extract_text()
            for page in pdf_reader.pages
            if page.extract_text() and page.extract_text().strip()
        ]
        return "\n\n".join(text)
    except Exception as e:
        st.error(f"Erro ao ler arquivo PDF: {e}")
        return None


def read_text_file(file) -> str:
    try:
        file.seek(0)
        raw = file.read()
        if isinstance(raw, bytes):
            return raw.decode("utf-8", errors="ignore")
        return str(raw)
    except Exception as e:
        st.error(f"Erro ao ler arquivo de texto: {e}")
        return None


def read_excel_file(file) -> str:
    try:
        file.seek(0)
        # Identifica a extensão do arquivo
        file_extension = file.name.split('.')[-1].lower()
        
        # Define o engine correto: xlrd para .xls e openpyxl para .xlsx
        engine = "openpyxl" if file_extension == "xlsx" else "xlrd"
        
        # Lê o arquivo Excel com o engine apropriado
        xls = pd.ExcelFile(file, engine=engine)
        
        blocos = []
        for sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name, dtype=str)
            df.fillna("", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df.dropna(axis=0, how="all", inplace=True)
            
            if df.empty:
                continue
                
            # Remove quebras de linha internas em células
            df = df.apply(lambda col: col.map(
                lambda v: re.sub(r"[\r\n]+", " ", str(v)).strip() if isinstance(v, str) else v
            ))
            
            csv_str = df.to_csv(index=False, sep=",", encoding="utf-8")
            blocos.append(f"=== ABA: {sheet_name} ===\n{csv_str}")
            
        if not blocos:
            st.warning("O arquivo Excel não contém dados nas abas.")
            return None
            
        return "\n\n".join(blocos)
    except Exception as e:
        st.error(f"Erro ao ler arquivo Excel (.{file_extension}): {e}")
        return None


def extract_text_from_file(uploaded_file) -> str:
    if uploaded_file is None:
        return None
    extension = uploaded_file.name.rsplit(".", 1)[-1].lower()
    uploaded_file.seek(0)
    readers = {
        "docx": read_word_file,
        "pdf": read_pdf_file,
        "txt": read_text_file,
        "csv": read_text_file,
        "xlsx": read_excel_file,
        "xls": read_excel_file,
    }
    reader = readers.get(extension)
    if reader:
        return reader(uploaded_file)
    st.error(f"Formato de arquivo não suportado: .{extension}")
    return None


# =============================================================================
# FUNÇÕES DE TRANSFORMAÇÃO DE ARQUIVOS
# =============================================================================

# =============================================================================
# CONSTANTES
# =============================================================================

# Padrão da linha de título informativo (descartada)
_RE_LINHA_INFORMATIVA = re.compile(
    r"CONTROLE DE EXAMES E PROCEDIMENTOS NA ENDOSCOPIA",
    re.IGNORECASE,
)

# Colunas canônicas do formato PRODUCAO com QTD (2026) — 13 campos
_COLUNAS_COM_QTD = [
    "QTD", "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# Colunas canônicas do formato PRODUCAO sem QTD (legado 2025) — 13 campos
# O primeiro campo é Data (sem QTD), e há EXAME REALIZADO 1 + EXAME REALIZADO 2
_COLUNAS_SEM_QTD = [
    "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento", "Procedimento2",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# Colunas canônicas do formato PRODUCAO 2025 NOVO (MAIO+ 2025) — 12 campos
# Sem QTD, sem EXAME REALIZADO 1/2, apenas EXAME REALIZADO único
_COLUNAS_2025_NOVO = [
    "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# =============================================================================
# MAPEAMENTOS DE COLUNAS → NOMES PADRONIZADOS
# =============================================================================

_MAP_PRODUCAO = {
    "qtd":                                   "QTD",
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",      # com ou sem espaço trailing
    "exame realizado 1":                     "Procedimento",
    "exame realizado 2":                     "Procedimento2",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adocional":                "ProcedimentosAdicionais",  # typo real do arquivo
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

# Mapeamento para formato 2025 NOVO (MAIO+ 2025) - sem QTD, sem EXAME REALIZADO 1/2
_MAP_PRODUCAO_2025_NOVO = {
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

# Mapeamento para formato 2026 (com QTD no início)
_MAP_PRODUCAO_2026 = {
    "qtd":                                   "QTD",
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",
    "procedimento":                          "Procedimento",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

_MAP_REPASSE = {
    "ds estabelecimento": "Estabelecimento",
    "cnpj estabelecimento": "CNPJ",
    "ds terceiro": "Terceiro",
    "ds status": "Status",
    "nr repasse terceiro": "NrRepasse",
    "tipo": "TipoItem",
    "nr atendimento": "NrAtendimento",
    "tp atend": "TipoAtendimento",
    "nr interno conta": "NrInternoConta",
    "paciente": "Paciente",
    "convenio": "Convenio",
    "convênio": "Convenio",
    "ds categoria": "Categoria",
    "cód item tuss": "CodigoTUSS",
    "cod item tuss": "CodigoTUSS",
    "ds procedimento": "Procedimento",
    "via unid med": "Via",
    "nm medico executor": "MedicoExecutor",
    "médico executor": "MedicoExecutor",
    "porcentagem": "Porcentagem",
    "ds funcao": "Funcao",
    "ds especialidade": "Especialidade",
    "qt procedimento": "QtProcedimento",
    "dt procedimento": "Data",
    "vl liberado": "ValorLiberado",
    "valor liberado": "ValorLiberado"
}

# Conjunto de todas as colunas canônicas do REPASSE (derivado automaticamente do mapa)
# Usado para garantir que todas as colunas estejam presentes após o processamento
_COLUNAS_CANONICAS_REPASSE = set(_MAP_REPASSE.values())

# =============================================================================
# FUNÇÕES AUXILIARES PRIVADAS
# =============================================================================

def _normalizar_coluna(col: str) -> str:
    return re.sub(r"\s+", " ", col.strip()).lower()


def _renomear_colunas(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    novo = {
        col: mapa[_normalizar_coluna(col)]
        for col in df.columns
        if _normalizar_coluna(col) in mapa
    }
    return df.rename(columns=novo)


def _detectar_formato_producao(df: pd.DataFrame, nome_aba: str = "") -> str:
    """
    Detecta o formato da planilha baseado nas colunas e nome da aba.
    Retorna: '2025_LEGADO' (JAN-ABR 2025), '2025_NOVO' (MAI+ 2025), '2026'
    """
    colunas_norm = [_normalizar_coluna(col) for col in df.columns]
    aba_upper = nome_aba.upper()
    
    # Formato 2025 LEGADO: tem "EXAME REALIZADO 1" e "EXAME REALIZADO 2"
    if "exame realizado 1" in colunas_norm or "exame realizado 2" in colunas_norm:
        return "2025_LEGADO"
    
    # Formato 2026: tem coluna QTD no início
    if "qtd" in colunas_norm:
        return "2026"
    
    # Formato 2025 NOVO (a partir de MAIO): sem QTD, sem EXAME REALIZADO 1/2
    # Tem apenas "EXAME REALIZADO" e "PROCEDIMENTOS ADICIONAIS"
    if "exame realizado" in colunas_norm and "procedimentos adicionais" in colunas_norm:
        return "2025_NOVO"
    
    # Fallback: tenta detectar pelo nome da aba
    if "2025" in aba_upper:
        # Se é JAN-ABR 2025, provavelmente é legado
        meses_legado = ["JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL"]
        if any(mes in aba_upper for mes in meses_legado):
            return "2025_LEGADO"
        return "2025_NOVO"
    
    return "2026"


def _padronizar_data(valor: str) -> str:
    """
    Converte qualquer formato de data reconhecível para DD/MM/AAAA.

    Formatos suportados (em ordem de tentativa):
      1. Serial numérico do Excel  — ex: 45344  → 22/02/2024
         Origem: arquivos de REPASSE exportados diretamente do Excel sem
         formatação de célula, onde a data fica como inteiro (dias desde
         30/12/1899, conforme convenção do Excel/Lotus).
         Faixa válida: 40000–50000 (aprox. 2009–2036) para evitar
         confundir com NrAtendimento ou outros números de 5 dígitos.
      2. ISO 8601  — ex: 2024-04-24 ou 2024-04-24 00:00:00
      3. Demais formatos reconhecidos pelo pandas (dayfirst=True)
    """
    if pd.isna(valor):
        return ""

    s = str(valor).strip()

    try:
        # 1. Serial numérico do Excel (5 dígitos na faixa 40000–50000)
        if re.match(r'^\d{5}$', s) and 40000 <= int(s) <= 50000:
            from datetime import date, timedelta
            dt = date(1899, 12, 30) + timedelta(days=int(s))
            return dt.strftime("%d/%m/%Y")

        # 2. ISO 8601 (YYYY-MM-DD ou YYYY-MM-DD HH:MM:SS)
        if re.match(r'^\d{4}-\d{2}-\d{2}', s):
            dt = pd.to_datetime(s, format='ISO8601')
            return dt.strftime("%d/%m/%Y")

        # 3. Outros formatos reconhecidos pelo pandas
        dt = pd.to_datetime(s, dayfirst=True)
        return dt.strftime("%d/%m/%Y")

    except Exception:
        return s


def _corrigir_data_malformada(data: str, aba_origem: str) -> str:
    """
    Corrige datas malformadas usando o mês da AbaOrigemDados como referência.
    Exemplos:
    - 18/032025 + ABA: MARÇO 2025 → 18/03/2025
    - 170/2025 + ABA: MAIO 2025 → 17/05/2025
    - 2706/2025 + ABA: JUNHO 2025 → 27/06/2025
    """
    if not data or not aba_origem:
        return data
    
    # Extrai mês e ano da AbaOrigemDados
    meses = {
        "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "MARCO": "03",
        "ABRIL": "04", "MAIO": "05", "JUNHO": "06",
        "JULHO": "07", "AGOSTO": "08", "SETEMBRO": "09",
        "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12"
    }
    
    mes_ref = None
    ano_ref = None
    aba_upper = aba_origem.upper()
    
    for nome_mes, num_mes in meses.items():
        if nome_mes in aba_upper:
            mes_ref = num_mes
            # Extrai ano (4 dígitos)
            match_ano = re.search(r'20\d{2}', aba_upper)
            if match_ano:
                ano_ref = match_ano.group()
            break
    
    if not mes_ref or not ano_ref:
        return data
    
    # Padrões de datas malformadas
    data_limpa = data.replace("/", "").replace("-", "").strip()
    
    # Padrão: 18032025 (sem separadores)
    if len(data_limpa) == 8 and data_limpa.isdigit():
        dia = data_limpa[:2]
        mes = data_limpa[2:4]
        ano = data_limpa[4:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    # Padrão: 170/2025 (dia com 3 dígitos)
    if re.match(r'^\d{3}/\d{4}$', data):
        dia = data[:2]  # Pega os 2 primeiros dígitos
        ano = data[-4:]
        return f"{dia}/{mes_ref}/{ano}"
    
    # Padrão: 2706/2025 (dia+mes sem separador)
    if re.match(r'^\d{4}/\d{4}$', data):
        dia = data[:2]
        mes = data[2:4]
        ano = data[-4:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    # Padrão: 17/042025 (ano sem separador)
    if re.match(r'^\d{2}/\d{6}$', data):
        dia = data[:2]
        mes = data[3:5]
        ano = data[5:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    return data


def _padronizar_valor(valor: str) -> str:
    """Remove R$, espaços e normaliza separadores decimais."""
    s = str(valor).strip() if valor else ""
    if not s:
        return ""
    s = re.sub(r"[R$\s]", "", s)
    if re.search(r"\d\.\d{3},", s):      # 1.234,56 → 1234.56
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")           # 664,61 → 664.61
    return re.sub(r"[,.]$", "", s)


def _identificar_tipo_arquivo(nome_aba: str, nome_arquivo: str = "") -> str:
    """
    Retorna 'REPASSE' ou 'PRODUCAO'.
    Prioridade: nome_aba → nome_arquivo → fallback PRODUCAO.
    """
    if "REPASSE" in nome_aba.upper():
        return "REPASSE"
    if "REPASSE" in nome_arquivo.upper():
        return "REPASSE"
    return "PRODUCAO"


def _detectar_tipo_por_cabecalho(linhas: list[str]) -> str:
    """
    Detecta o tipo (PRODUCAO ou REPASSE) inspecionando as primeiras linhas.
    Usa colunas exclusivas do REPASSE como indicadores.
    """
    _COLS_REPASSE = {
        "ds estabelecimento", "ds terceiro", "nr repasse terceiro",
        "ds procedimento", "nm medico executor", "dt procedimento",
        "vl liberado", "cód item tuss", "ds categoria",
    }
    for linha in linhas[:3]:
        cols = {c.strip().lower() for c in linha.split(",")}
        if cols & _COLS_REPASSE:
            return "REPASSE"
    return "PRODUCAO"


def _linha_e_valida(linha: str) -> bool:
    """
    Descarta:
    - linhas em branco ou só vírgulas/espaços
    - linhas com apenas 1 valor não vazio (ex: "429,,,,,,,,,,,")
    - linha de título informativo ("CONTROLE DE EXAMES...")
    - linha com colunas todas 'Unnamed: N'
    """
    s = linha.strip()
    if not s or re.match(r"^[,\s]*$", s):
        return False
    nao_vazios = [p.strip() for p in s.split(",") if p.strip()]
    if len(nao_vazios) <= 1:
        return False
    if _RE_LINHA_INFORMATIVA.search(s):
        return False
    if all(re.match(r"^unnamed\s*:\s*\d+$", v, re.I) for v in nao_vazios):
        return False
    
    return True


def _detectar_cabecalho_producao(linhas: list[str]) -> tuple[str, int]:
    """
    Inspeciona as primeiras linhas válidas e retorna:
      ("COM_QTD", idx)   — cabeçalho explícito com a coluna QTD
      ("SEM_QTD", idx)   — cabeçalho explícito sem QTD (formato 2025)
      ("SEM_HEADER", 0)  — sem cabeçalho; dados começam na linha 0

    idx = índice em `linhas` onde está o cabeçalho (ou 0 se não houver).
    """
    for idx, linha in enumerate(linhas[:5]):
        upper = linha.upper()
        campos = [c.strip() for c in linha.split(",")]

        # Cabeçalho 2026 com QTD
        if "QTD" in campos[0].upper() and "DATA" in upper:
            return "COM_QTD", idx

        # Cabeçalho 2025 sem QTD — começa com DATA
        if campos[0].strip().upper() in ("DATA",) and "NOME DO PACIENTE" in upper:
            return "SEM_QTD", idx

    return "SEM_HEADER", 0


def _atribuir_colunas(df: pd.DataFrame, formato: str, nome_aba: str = "") -> pd.DataFrame:
    """
    Atribui cabeçalho canônico a DataFrames sem cabeçalho explícito.
    """
    n = len(df.columns)
    
    # Detecta qual base usar baseado no número de colunas e nome da aba
    if n == 12:
        # 12 colunas = formato 2025 NOVO (MAIO+ 2025)
        base = _COLUNAS_2025_NOVO
    elif n == 13:
        # 13 colunas: pode ser COM_QTD (2026) ou SEM_QTD (2025 LEGADO)
        # Verifica se primeira coluna parece QTD
        amostra = df.iloc[:, 0].dropna().head(5).tolist()
        primeiro_parece_qtd = all(
            re.match(r"^\d*$", str(v).strip()) for v in amostra if str(v).strip()
        )
        base = _COLUNAS_COM_QTD if primeiro_parece_qtd else _COLUNAS_SEM_QTD
    else:
        # Fallback: tenta detectar pelo nome da aba
        if "2025" in nome_aba.upper():
            meses_legado = ["JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL"]
            if any(mes in nome_aba.upper() for mes in meses_legado):
                base = _COLUNAS_SEM_QTD
            else:
                base = _COLUNAS_2025_NOVO
        else:
            base = _COLUNAS_COM_QTD

    df.columns = (base + [f"Extra_{k}" for k in range(n - len(base))])[:n] if n <= len(base) else \
                 base + [f"Extra_{k}" for k in range(n - len(base))]
    return df


# =============================================================================
# PROCESSADORES POR TIPO
# =============================================================================

def _processar_aba_producao(df: pd.DataFrame, nome_aba: str) -> pd.DataFrame:
    """
    Normaliza um DataFrame de PRODUCAO:
    - Detecta formato (2025_LEGADO, 2025_NOVO, 2026)
    - Renomeia colunas pelo mapa apropriado
    - Garante QTD (0 se ausente), Data, Paciente, Procedimento
    - Gera linhas extras para 'Procedimento2' (EXAME REALIZADO 2 das abas 2025 LEGADO)
    - Adiciona TipoArquivo = 'PRODUCAO' e AbaOrigemDados
    - Descarta linhas sem Paciente válido
    """
    # Detecta formato baseado nas colunas e nome da aba
    formato = _detectar_formato_producao(df, nome_aba)
    
    # Seleciona mapa apropriado
    if formato == "2025_LEGADO":
        mapa = _MAP_PRODUCAO
    elif formato == "2025_NOVO":
        mapa = _MAP_PRODUCAO_2025_NOVO
    else:  # 2026
        mapa = _MAP_PRODUCAO_2026
    
    # CORREÇÃO: Remove primeira coluna se for Unnamed e formato não é 2026
    # Isso acontece quando pandas lê cabeçalho sem QTD mas dados têm coluna extra
    if formato != "2026" and len(df.columns) > 0:
        primeira_col = str(df.columns[0]).lower()
        if "unnamed" in primeira_col or primeira_col.strip() in ("", "nan"):
            df = df.iloc[:, 1:]  # Remove primeira coluna
    
    df = _renomear_colunas(df, mapa)
    
    # CORREÇÃO: Limpa valores "Unnamed: N" dos dados (problema do Excel)
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: "" if str(v).strip().upper().startswith("UNNAMED:") else v
        )

    for col in ("QTD", "Data", "Paciente", "Procedimento", "Procedimento2",
                "Convenio", "Origem", "NrAtendimento", "MedicoExecutor",
                "LocalSetor", "Sala", "Carater", "Observacao",
                "ProcedimentosAdicionais"):
        if col not in df.columns:
            df[col] = ""

    # Normaliza QTD
    df["QTD"] = df["QTD"].apply(
        lambda v: "0" if str(v).strip() in ("", "nan", "NaN") else str(v).strip()
    )

    # Primeiro aplica padronização básica
    df["Data"] = df["Data"].apply(_padronizar_data)
    
    # Depois corrige datas malformadas usando AbaOrigemDados
    df["Data"] = df.apply(
        lambda row: _corrigir_data_malformada(row["Data"], nome_aba),
        axis=1
    )
    
    # Tenta padronizar novamente após correção
    df["Data"] = df["Data"].apply(_padronizar_data)
    
    df["Paciente"]     = df["Paciente"].apply(
        lambda v: re.sub(r"\s+", " ", str(v).strip()).upper()
    )
    df["Procedimento"] = df["Procedimento"].apply(
        lambda v: str(v).strip().upper() if str(v).strip() not in ("", "-", "nan") else ""
    )

    linhas_saida = []
    for _, row in df.iterrows():
        paciente = str(row.get("Paciente", "")).strip()
        # Descarta linhas de formatação sem paciente real
        if not paciente or paciente in ("NAN", ""):
            continue

        linha_principal = row.copy()
        linha_principal["TipoArquivo"]    = "PRODUCAO"
        linha_principal["AbaOrigemDados"] = f"ABA: {nome_aba}"
        linhas_saida.append(linha_principal)

        # Trata Procedimento2 (EXAME REALIZADO 2 das abas 2025 LEGADO) como linha extra
        proc2 = str(row.get("Procedimento2", "")).strip().upper()
        if proc2 and proc2 not in ("", "-", "NAN", "NONE"):
            extra = row.copy()
            extra["Procedimento"]    = proc2
            extra["Procedimento2"]   = ""
            extra["Observacao"]      = "PROCEDIMENTO_ADICIONAL"
            extra["TipoArquivo"]     = "PRODUCAO"
            extra["AbaOrigemDados"]  = f"ABA: {nome_aba}"
            extra["QTD"]             = "0"
            linhas_saida.append(extra)

    df_saida = pd.DataFrame(linhas_saida)
    df_saida.drop(columns=["Procedimento2"], inplace=True, errors="ignore")
    # Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    lixo = [c for c in df_saida.columns
            if re.match(r"^Extra_\d+$", str(c)) or re.search(r"\.\d+$", str(c))]
    df_saida.drop(columns=lixo, inplace=True, errors="ignore")
    return df_saida


def _processar_aba_repasse(df: pd.DataFrame, nome_aba: str) -> pd.DataFrame:
    """
    Normaliza um DataFrame de REPASSE:
    - Renomeia colunas pelo mapa
    - Limpa valores "Unnamed: N" gerados pelo pandas ao ler Excel
    - Garante que TODAS as colunas canonicas do _MAP_REPASSE existam (vazias se ausentes)
    - Padroniza Data, Paciente e ValorLiberado
    - Adiciona TipoArquivo = 'REPASSE' e AbaOrigemDados
    - Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    """
    df = _renomear_colunas(df, _MAP_REPASSE)

    # Limpa valores "Unnamed: N" nos dados (problema do Excel com colunas sem cabecalho)
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: "" if str(v).strip().upper().startswith("UNNAMED:") else v
        )

    # Garante TODAS as colunas canonicas do _MAP_REPASSE, nao apenas um subconjunto.
    # Isso evita que colunas ausentes no arquivo fisico (ex: NrAtendimento) sejam
    # omitidas do CSV de saida. Derivado automaticamente do mapa - sem listas manuais.
    for col in _COLUNAS_CANONICAS_REPASSE:
        if col not in df.columns:
            df[col] = ""

    df["Data"]          = df["Data"].apply(_padronizar_data)
    df["Paciente"]      = df["Paciente"].apply(
        lambda v: re.sub(r"\s+", " ", str(v).strip()).upper()
    )
    df["ValorLiberado"] = df["ValorLiberado"].apply(_padronizar_valor)
    df["TipoArquivo"]   = "REPASSE"
    df["AbaOrigemDados"] = f"ABA: {nome_aba}"

    # Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    # Alinhado com o mesmo tratamento ja feito em _processar_aba_producao
    lixo = [c for c in df.columns
            if re.match(r"^Extra_\d+$", str(c)) or re.search(r"\.\d+$", str(c))]
    df.drop(columns=lixo, inplace=True, errors="ignore")
    return df


# =============================================================================
# PROCESSAMENTO DE UM BLOCO TEXTO (UMA ABA)
# =============================================================================

def _processar_bloco_texto(
    dados_aba: str,
    nome_aba: str,
    nome_arquivo: str,
) -> str:
    """Filtra, detecta formato e processa o texto de uma única aba."""

    # FIX defensivo: colapsa \n internos em campos entre aspas (origem: CSV/TXT)
    dados_aba = re.sub(
        r'"([^"]*)"',
        lambda m: '"' + m.group(1).replace('\n', ' ').replace('\r', '') + '"',
        dados_aba,
    )
    # ── 1. Filtrar linhas inválidas ──────────────────────────────────────────
    linhas_validas = [l for l in dados_aba.splitlines() if _linha_e_valida(l)]
    if not linhas_validas:
        return ""

    tipo = _identificar_tipo_arquivo(nome_aba, nome_arquivo)
    if tipo == "PRODUCAO":
        tipo = _detectar_tipo_por_cabecalho(linhas_validas)

    # ── 2. REPASSE — cabeçalho sempre explícito ──────────────────────────────
    if tipo == "REPASSE":
        try:
            df = pd.read_csv(io.StringIO("\n".join(linhas_validas)), dtype=str)
            df.fillna("", inplace=True)
            df_proc = _processar_aba_repasse(df, nome_aba)
            return df_proc.to_csv(index=False, encoding="utf-8")
        except Exception as exc:
            logger.warning(f"Erro REPASSE '{nome_aba}': {exc}")
            return ""

    # ── 3. PRODUCAO — detecta formato ────────────────────────────────────────
    formato, idx_hdr = _detectar_cabecalho_producao(linhas_validas)

    if formato in ("COM_QTD", "SEM_QTD"):
        # Cabeçalho explícito na linha idx_hdr; descarta linhas anteriores
        linhas_csv = linhas_validas[idx_hdr:]
        try:
            df = pd.read_csv(io.StringIO("\n".join(linhas_csv)), dtype=str)
            df.fillna("", inplace=True)
        except Exception as exc:
            logger.warning(f"Erro PRODUCAO com header '{nome_aba}': {exc}")
            return ""
    else:
        # SEM_HEADER: nenhum cabeçalho encontrado → aplica cabeçalho canônico
        try:
            df = pd.read_csv(
                io.StringIO("\n".join(linhas_validas)),
                header=None, dtype=str,
            )
            df.fillna("", inplace=True)
            df = _atribuir_colunas(df, formato, nome_aba)
        except Exception as exc:
            logger.warning(f"Erro PRODUCAO sem header '{nome_aba}': {exc}")
            return ""

    df_proc = _processar_aba_producao(df, nome_aba)
    return df_proc.to_csv(index=False, encoding="utf-8")


def _consolidar_blocos(blocos_csv: list[str]) -> str:
    """Une múltiplos CSVs em um único, com cabeçalho único na primeira linha."""
    dfs = []
    for bloco in blocos_csv:
        bloco = bloco.strip()
        if not bloco:
            continue
        try:
            df = pd.read_csv(io.StringIO(bloco), dtype=str)
            df.fillna("", inplace=True)
            dfs.append(df)
        except Exception:
            pass
    if not dfs:
        return ""
    df_final = pd.concat(dfs, ignore_index=True, sort=False)
    df_final.fillna("", inplace=True)
    return df_final.to_csv(index=False, encoding="utf-8")


# =============================================================================
# FUNÇÃO PÚBLICA PRINCIPAL
# =============================================================================

def transformar_csv_arquivo(
    conteudo_texto: str,
    nome_arquivo: str = "",
    desmembrar_procedimentos_adicionais: bool = False,
) -> str:
    """
    Transforma o conteúdo extraído de um arquivo de endoscopia (PRODUCAO ou REPASSE)
    em um CSV padronizado puro, sem linhas vazias ou texto explicativo.

    Regras aplicadas:
    - Remove linhas completamente vazias, com só vírgulas (ex: "429,,,,,,,,,,,,")
      ou com apenas um valor não vazio
    - Remove linhas informativas de título ("CONTROLE DE EXAMES E PROCEDIMENTOS
      NA ENDOSCOPIA / MÊS ANO,Unnamed: 1,Unnamed: 2,...")
    - Extrai o nome da aba do marcador "=== ABA: <nome> ===" e insere na coluna
      AbaOrigemDados (ex: "ABA: JANEIRO 2026")
    - Primeira linha do CSV = cabeçalho obrigatório
    - Data padronizada para DD/MM/AAAA
    - Paciente em MAIÚSCULAS sem espaços extras
    - TipoArquivo = 'PRODUCAO' ou 'REPASSE' (identificado pelo nome da aba/arquivo)
    - Coluna QTD incluída; preenchida com "0" quando ausente na aba de origem
    - Abas com formato legado 2025 (EXAME REALIZADO 1 / EXAME REALIZADO 2) têm
      "EXAME REALIZADO 2" tratado como linha extra de procedimento
    - Todas as colunas canônicas do _MAP_REPASSE são sempre incluídas no CSV de REPASSE,
      mesmo que o arquivo físico não as contenha (ficam vazias)

    Args:
        conteudo_texto:   Texto extraído por extract_text_from_file.
                          Para Excel: contém blocos "=== ABA: <nome> ===" separando abas.
        nome_arquivo:     Nome original do arquivo (fallback para identificar tipo).
        desmembrar_procedimentos_adicionais:
                          Se True E TipoArquivo for PRODUCAO, chama
                          desmembrar_procedimentos_adicionais_csv_arquivo antes de retornar.

    Returns:
        CSV puro (str) com cabeçalho na primeira linha, ou "" em caso de falha total.
    """
    blocos_csv: list[str] = []

    # Divide por marcadores de aba (formato gerado por read_excel_file)
    partes = re.split(r"(===\s*ABA:\s*.+?===)", conteudo_texto)

    if len(partes) > 1:
        it = iter(partes[1:])
        for marcador, dados_aba in zip(it, it):
            nome_aba = re.sub(r"===\s*ABA:\s*|===", "", marcador).strip()
            dados_aba = dados_aba.strip()
            if not dados_aba:
                continue
            try:
                bloco = _processar_bloco_texto(dados_aba, nome_aba, nome_arquivo)
                if bloco:
                    blocos_csv.append(bloco)
            except Exception as exc:
                logger.warning(f"Falha ao processar aba '{nome_aba}': {exc}")
    else:
        # Arquivo CSV/TXT simples sem marcador de aba
        nome_aba = re.sub(r"\.[^.]+$", "", nome_arquivo)
        try:
            bloco = _processar_bloco_texto(conteudo_texto.strip(), nome_aba, nome_arquivo)
            if bloco:
                blocos_csv.append(bloco)
        except Exception as exc:
            logger.error(f"Falha ao processar '{nome_arquivo}': {exc}")
            return ""

    if not blocos_csv:
        return ""

    resultado = _consolidar_blocos(blocos_csv)

    # Desmembramento opcional (só PRODUCAO)
    if desmembrar_procedimentos_adicionais and _identificar_tipo_arquivo(nome_arquivo) == "PRODUCAO":
        resultado = desmembrar_procedimentos_adicionais_csv_arquivo(resultado)

    return resultado


# =============================================================================
# FUNÇÃO PÚBLICA: DESMEMBRAMENTO DE PROCEDIMENTOS ADICIONAIS
# =============================================================================

def desmembrar_procedimentos_adicionais_csv_arquivo(csv_texto: str) -> str:
    """
    Recebe um CSV de PRODUCAO já padronizado e desmembra a coluna
    'ProcedimentosAdicionais' em linhas extras.

    Regras:
    - Para cada valor em 'ProcedimentosAdicionais', gera uma LINHA EXTRA com:
        * Procedimento  = nome do procedimento adicional (MAIÚSCULAS)
        * Observacao    = "PROCEDIMENTO_ADICIONAL"
        * QTD           = "0"
        * demais campos = replicados da linha principal
    - A linha principal mantém sua Observacao original (vazia ou com valor já existente).
    - Não altera linhas de REPASSE (TipoArquivo != 'PRODUCAO').
    - Remove a coluna 'ProcedimentosAdicionais' do CSV final.
    - Delimitadores suportados: ponto-e-vírgula, '+', quebra de linha.
      Vírgulas são delimitadoras apenas quando não fazem parte de número decimal.

    Args:
        csv_texto: CSV puro (string) com cabeçalho na primeira linha.

    Returns:
        CSV puro (string) com as linhas extras já inseridas e sem
        a coluna 'ProcedimentosAdicionais'.
    """
    if not csv_texto or not csv_texto.strip():
        return csv_texto

    try:
        df = pd.read_csv(io.StringIO(csv_texto.strip()), dtype=str)
        df.fillna("", inplace=True)
    except Exception as exc:
        logger.warning(f"desmembrar_procedimentos_adicionais_csv_arquivo: erro ao ler CSV: {exc}")
        return csv_texto

    if "ProcedimentosAdicionais" not in df.columns:
        return csv_texto

    linhas_saida = []
    for _, row in df.iterrows():
        linhas_saida.append(row.copy())

        # Só desmembra linhas de PRODUCAO
        if str(row.get("TipoArquivo", "")).strip().upper() != "PRODUCAO":
            continue

        proc_adicionais = str(row.get("ProcedimentosAdicionais", "")).strip()
        if proc_adicionais.upper() in ("", "NAN", "NONE", "NAN", "-"):
            continue

        # Delimitadores: ; | + | quebra de linha
        # Vírgula apenas quando não separa dígitos (evita quebrar "1,234")
        separados = re.split(r"[;+\n\r]+|(?<!\d),(?!\d)", proc_adicionais)

        for proc in separados:
            proc = proc.strip().upper()
            if not proc or proc in ("-", ""):
                continue
            extra = row.copy()
            extra["Procedimento"]            = proc
            extra["ProcedimentosAdicionais"] = ""
            extra["Observacao"]              = "PROCEDIMENTO_ADICIONAL"
            extra["QTD"]                     = "0"
            linhas_saida.append(extra)

    df_saida = pd.DataFrame(linhas_saida)
    df_saida.drop(columns=["ProcedimentosAdicionais"], inplace=True, errors="ignore")
    df_saida.fillna("", inplace=True)
    return df_saida.to_csv(index=False, encoding="utf-8")

###############################################################################


# =============================================================================
# FUNÇÕES DE CORRELAÇÃO LOCAL
# =============================================================================

from datetime import datetime, timedelta
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Optional

# Dicionário de sinônimos e variações de procedimentos
SINONIMOS_PROCEDIMENTOS = {
    # Colonoscopia e variações
    "COLONO": ["COLONOSCOPIA", "COLONOSCOPIA INCLUI", "RETOSSIGMOIDOSCOPIA"],
    "COLONOSCOPIA": ["COLONO", "COLONOSCOPIA INCLUI", "RETOSSIGMOIDOSCOPIA"],
    "COLONOCOPIA": ["COLONO", "COLONOSCOPIA"],  # Typo comum
    "RETOSSIGMOIDOSCOPIA": ["COLONOSCOPIA", "COLONO"],
    
    # Endoscopia Digestiva Alta
    "ENDOSCOPIA": ["EDA", "ENDOSCOPIA DIGESTIVA ALTA", "ESOFAGOGASTRODUODENOSCOPIA", "ESOFAGO GASTRO DUODENOSCOPIA"],
    "EDA": ["ENDOSCOPIA", "ENDOSCOPIA DIGESTIVA ALTA", "ESOFAGOGASTRODUODENOSCOPIA"],
    "ENDOCOSPIA": ["ENDOSCOPIA"],  # Typo
    "ENDDOSCOPIA": ["ENDOSCOPIA"],  # Typo
    "ESOFAGOGASTRODUODENOSCOPIA": ["ENDOSCOPIA", "EDA"],
    "ESOFAGO GASTRO DUODENOSCOPIA": ["ENDOSCOPIA", "EDA"],
    
    # Ecoendoscopia
    "ECOENDOSCOPIA": ["ECOEDA", "ULTRASSOM ENDOSCOPICO"],
    "ECOEDA": ["ECOENDOSCOPIA", "ULTRASSOM ENDOSCOPICO"],
    
    # Teste de Urease / H. Pylori
    "TESTE": ["PESQUISA", "EXAME"],
    "UREASE": ["H PYLORI", "HELICOBACTER", "HELICOBACTER PYLORI"],
    "TESTE DE UREASE": ["PESQUISA H PYLORI", "TESTE UREASE", "UREASE", "H PYLORI", "HELICOBACTER"],
    "TESTE DA UREASE": ["TESTE DE UREASE", "PESQUISA H PYLORI", "HELICOBACTER"],
    "HP": ["H PYLORI", "HELICOBACTER"],
    "HPYLORI": ["H PYLORI", "HELICOBACTER"],
    "HELICOBACTER": ["H PYLORI", "UREASE", "HP"],
    "HELICOBACTER PYLORI": ["H PYLORI", "UREASE", "HP"],
    
    # Anátomo Patológico / Biópsia / Citologia
    "ANATOMO": ["BIOPSIA", "CITOLOGIA", "AP"],
    "ANATOMIA": ["BIOPSIA", "CITOLOGIA", "AP"],
    "ANATOMIA PATOLOGICA": ["BIOPSIA", "CITOLOGIA", "AP", "ANÁTOMO PATOLÓGICO"],
    "ANATOMO PATOLOGICA": ["BIOPSIA", "CITOLOGIA", "AP", "ANÁTOMO PATOLÓGICO"],
    "ANATOMO PATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO", "ANATOMOPATOLOGICO"],
    "ANÁTOMO PATOLÓGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANATOMO PATOLOGICO", "ANATOMOPATOLOGICO"],
    "ANÁTOMO PATLÓGICO": ["ANÁTOMO PATOLÓGICO", "BIOPSIA", "AP"],  # Typo
    "ANATOMOPATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANATOMO PATOLOGICO"],
    "ANÁTOMOPATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO"],
    "AP": ["ANATOMO PATOLOGICO", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO", "ANATOMOPATOLOGICO"],
    "BIOPSIA": ["ANATOMO PATOLOGICO", "AP", "CITOLOGIA", "BIOPSIAS"],
    "BIOPSIAS": ["BIOPSIA", "ANATOMO PATOLOGICO", "AP"],
    "BIOPSIA SERIADO": ["BIOPSIA", "ANÁTOMO PATOLÓGICO", "AP"],
    "CITOLOGIA": ["BIOPSIA", "ANATOMO PATOLOGICO", "AP"],
    
    # Polipectomia
    "POLIPECTOMIA": ["POLIPO", "RESSECCAO POLIPO", "RETIRADA POLIPO"],
    "POLIPO": ["POLIPECTOMIA", "RESSECCAO POLIPO"],
    "PÓLIPO": ["POLIPECTOMIA", "POLIPO"],
    "POLIOECTOMIA": ["POLIPECTOMIA"],  # Typo
    
    # Hemostasia
    "HEMOSTASIA": ["HEMOSTASE", "CONTROLE SANGRAMENTO", "HEMOSTASE ENDOSCOPICA"],
    "HEMOSTASE": ["HEMOSTASIA", "CONTROLE SANGRAMENTO"],
    
    # Mucosectomia
    "MUCOSECTOMIA": ["RESSECCAO MUCOSA", "RESSECCAO ENDOSCOPICA"],
    
    # Tatuagem
    "TATUAGEM": ["MARCACAO", "TATUAGEM ENDOSCOPICA"],
    
    # Biópsia Hepática
    "BIOPSIA HEPATICA": ["BIOPSIA", "HEPATICA"],
    "HEPATICA": ["BIOPSIA HEPATICA", "FIGADO"],
    
    # CPRE / Colangiopancreatografia / Papilotomia
    "CPRE": ["COLANGIOPANCREATOGRAFIA", "COLANGIOPANCREATOGRAFIA RETROGRADA", "PAPILOTOMIA"],
    "COLANGIOPANCREATOGRAFIA": ["CPRE", "PAPILOTOMIA", "COLANGIOPANCREATOGRAFIA RETROGRADA"],
    "COLANGIOPANCREATOGRAFIA RETROGRADA": ["CPRE", "COLANGIOPANCREATOGRAFIA", "PAPILOTOMIA"],
    "PAPILOTOMIA": ["CPRE", "COLANGIOPANCREATOGRAFIA"],
    "PAPILOTOMIA ENDOSCOPICA": ["CPRE", "PAPILOTOMIA"],
    
    # Dilatação
    "DILATACAO": ["DILATAÇÃO", "DILATACAO PNEUMATICA"],
    "DILATAÇÃO": ["DILATACAO", "DILATACAO PNEUMATICA"],
    
    # Gastrostomia
    "GASTROSTOMIA": ["GTT", "GASTROSTOMIA ENDOSCOPICA"],
    "GASTROSTOMIA ENDOSCOPICA": ["GASTROSTOMIA", "GTT"],
    "GTT": ["GASTROSTOMIA"],
    
    # Passagem de Sonda
    "SONDA": ["PASSAGEM SONDA", "SONDA NASO ENTERAL", "SONDA NASOENTERAL"],
    "PASSAGEM DE SONDA": ["SONDA", "SONDA NASO ENTERAL", "PASSAGEM SONDAS"],
    "PASSAGEM DE SONDAS": ["PASSAGEM DE SONDA", "SONDA"],
    "SONDA NASO ENTERAL": ["SONDA", "PASSAGEM SONDA"],
    
    # Anuscopia / Retoscopia
    "ANUSCOPIA": ["RETOSCOPIA", "EXAME ANORRETAL"],
    "RETOSCOPIA": ["ANUSCOPIA"],
    
    # Prótese
    "PROTESE": ["COLOCACAO PROTESE", "IMPLANTE PROTESE"],
    "PRÓTESE": ["PROTESE", "COLOCACAO PROTESE"],
    "COLOCACAO DE PROTESE": ["PROTESE", "IMPLANTE PROTESE"],
    "COLOCACAO PROTESE": ["PROTESE", "COLOCACAO DE PROTESE"],
    
    # Visita Hospitalar
    "VISITA": ["VISITA HOSPITALAR", "VISITA PACIENTE INTERNADO", "HOSPITALAR"],
    "VISITA HOSPITALAR": ["VISITA", "VISITA PACIENTE INTERNADO", "HOSPITALAR"],
    "HOSPITALAR": ["VISITA", "VISITA HOSPITALAR"],
    
    # Cromoscopia
    "CROMOSCOPIA": ["CROMOSCOPIA ENDOSCOPICA", "COLORACAO"],
    
    # Esôfago
    "ESOFAGO": ["ESOFAGICO", "ESOFAGICA"],
    "ESOFAGICO": ["ESOFAGO"],
    
    # Cálculo / Coledociano
    "CALCULO": ["CALCULOSE", "PEDRA"],
    "COLEDOCIANO": ["COLECOCO", "BILIAR"],
    
    # Drenagem
    "DRENAGEM": ["DRENAGEM BILIAR", "DESCOMPRESSAO"],
    "DESCOMPRESSAO": ["DRENAGEM"],
    "DESCOMPRESSAO COLONICA": ["DESCOMPRESSAO", "COLONICA"],
    "COLONICA": ["COLON", "COLONOSCOPIA"],
    
    # Ostomia
    "OSTOMIA": ["GASTROSTOMIA", "COLOSTOMIA"],
    "TROCA DE GTT": ["GTT", "GASTROSTOMIA"],
}

# Pares anatomicamente incompatíveis: EDA (trato alto) ≠ Colonoscopia (trato baixo).
# Usados para sinalizar matches via SequenceMatcher acidental (ex: ENDOSCOPIA↔COLONOSCOPIA
# compartilham "OSCOPIA" como substring, gerando sim=0.73 acima do limiar 0.65).
_PARES_ANATOMICAMENTE_DIVERGENTES: list[frozenset] = [
    frozenset({"ENDOSCOPIA", "COLONOSCOPIA"}),
    frozenset({"ENDOSCOPIA", "COLONO"}),
    frozenset({"ENDOSCOPIA", "RETOSSIGMOIDOSCOPIA"}),
    frozenset({"EDA", "COLONOSCOPIA"}),
    frozenset({"EDA", "COLONO"}),
    frozenset({"EDA", "RETOSSIGMOIDOSCOPIA"}),
]

# Palavras-chave de procedimentos companion (Urease/Helicobacter standalone).
_KEYWORDS_COMPANION_PROC = {"UREASE", "HELICOBACTER"}
# Palavras-chave de procedimentos principais — excluem uma linha do perfil companion.
_KEYWORDS_PRINCIPAL_PROC = {
    "ENDOSCOPIA", "COLONOSCOPIA", "COLONO", "ECOENDOSCOPIA",
    "GASTROSTOMIA", "MUCOSECTOMIA", "CPRE", "COLANGIOPANCREATOGRAFIA",
    "PAPILOTOMIA", "BRONCOSCOPIA", "RETOSSIGMOIDOSCOPIA",
}


def _sao_anatomicamente_divergentes(proc1: str, proc2: str) -> bool:
    """
    Retorna True quando os dois procedimentos pertencem a grupos anatômicos
    reconhecidamente distintos (ex: ENDOSCOPIA/EDA vs COLONOSCOPIA/trato baixo).
    Não bloqueia o match — apenas sinaliza para revisão humana via status.
    """
    p1 = _normalizar_procedimento(proc1)
    p2 = _normalizar_procedimento(proc2)
    tokens1 = set(p1.split())
    tokens2 = set(p2.split())
    for par in _PARES_ANATOMICAMENTE_DIVERGENTES:
        t1, t2 = tuple(par)
        if (t1 in tokens1 and t2 in tokens2) or (t2 in tokens1 and t1 in tokens2):
            return True
    return False


def _e_procedimento_companion(proc: str) -> bool:
    """
    Retorna True para procedimentos reconhecidamente acompanhantes/adicionais
    (ex: Teste de Urease standalone) que na PRODUCAO costumam estar em
    ProcedimentosAdicionais. Exclui procedimentos compostos que já contêm o
    exame principal no nome (ex: 'EDA Com Biópsia E Teste De Urease').
    """
    p_norm = _normalizar_procedimento(proc)
    tokens = set(p_norm.split())
    return bool(tokens & _KEYWORDS_COMPANION_PROC) and not bool(tokens & _KEYWORDS_PRINCIPAL_PROC)


def _normalizar_procedimento(proc: str) -> str:
    """Normaliza nome de procedimento para comparação semântica."""
    if not proc or str(proc).strip() in ("", "nan", "NaN"):
        return ""
    s = str(proc).upper().strip()
    # Remove acentos
    s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    s = s.replace('Â', 'A').replace('Ê', 'E').replace('Ô', 'O')
    s = s.replace('Ã', 'A').replace('Õ', 'O')
    s = s.replace('Ç', 'C')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Remove plural (S no final de palavras)
    s = re.sub(r'\bBIOPSIAS\b', 'BIOPSIA', s)
    s = re.sub(r'\bPOLIPOS\b', 'POLIPO', s)
    s = re.sub(r'\bSONDAS\b', 'SONDA', s)
    # Remove separadores comuns em procedimentos compostos
    s = re.sub(r'\s*[/+]\s*', ' ', s)  # Remove / e +
    return s

def _expandir_abreviacao(proc: str) -> str:
    """Expande abreviações comuns de procedimentos."""
    ABREVIACOES = {
        "COLONO": "COLONOSCOPIA",
        "EDA": "ENDOSCOPIA DIGESTIVA ALTA",
        "RETO": "RETOSSIGMOIDOSCOPIA",
        "RETOSSIGMOIDO": "RETOSSIGMOIDOSCOPIA",
        "POLIPECTOMIA": "POLIPECTOMIA",
        "LIGADURA": "LIGADURA ELASTICA",
        "DILATACAO": "DILATACAO ESOFAGICA",
        "ESCLEROSE": "ESCLEROTERAPIA",
        "MUCOSECTOMIA": "MUCOSECTOMIA",
        "HEMOSTASIA": "HEMOSTASIA ENDOSCOPICA",
    }
    proc_norm = _normalizar_procedimento(proc)
    
    # Verifica match exato
    if proc_norm in ABREVIACOES:
        return ABREVIACOES[proc_norm]
    
    # Verifica se começa com abreviação
    for abrev, expandido in ABREVIACOES.items():
        if proc_norm.startswith(abrev + " ") or proc_norm == abrev:
            return proc_norm.replace(abrev, expandido, 1)
    
    return proc_norm

def _match_palavra_contida(proc1: str, proc2: str) -> bool:
    """Verifica se um procedimento está contido no outro ou compartilha palavra-chave."""
    p1 = _normalizar_procedimento(proc1).strip()
    p2 = _normalizar_procedimento(proc2).strip()
    
    if not p1 or not p2:
        return False
    
    # Verifica contenção direta
    if p1 in p2 or p2 in p1:
        return True
    
    # Verifica se primeira palavra significativa é igual (mínimo 4 caracteres)
    palavras1 = [p for p in p1.split() if len(p) >= 4]
    palavras2 = [p for p in p2.split() if len(p) >= 4]
    
    if palavras1 and palavras2:
        # Verifica se primeira palavra significativa é igual
        if palavras1[0] == palavras2[0]:
            return True
        # Verifica se alguma palavra significativa está contida
        for palavra in palavras1:
            if len(palavra) >= 6 and any(palavra in p2_word or p2_word in palavra for p2_word in palavras2):
                return True
    
    return False

def _verificar_sinonimo(proc1: str, proc2: str) -> bool:
    """Verifica se dois procedimentos são sinônimos conhecidos."""
    p1_norm = _normalizar_procedimento(proc1)
    p2_norm = _normalizar_procedimento(proc2)
    
    # Verifica match direto no dicionário
    for chave, sinonimos in SINONIMOS_PROCEDIMENTOS.items():
        if chave in p1_norm:
            for sin in sinonimos:
                if sin in p2_norm:
                    return True
        if chave in p2_norm:
            for sin in sinonimos:
                if sin in p1_norm:
                    return True
    
    # Para procedimentos compostos (ex: "ANATOMO + POLIPECTOMIA")
    # Verifica se todas as palavras-chave de um estão no outro
    palavras1 = set(p1_norm.split())
    palavras2 = set(p2_norm.split())
    
    # Se tem pelo menos 2 palavras em comum e são palavras-chave
    palavras_chave = {'BIOPSIA', 'POLIPECTOMIA', 'POLIPO', 'ANATOMO', 'TESTE', 'UREASE', 
                      'MUCOSECTOMIA', 'HEMOSTASIA', 'ENDOSCOPIA', 'COLONOSCOPIA', 'COLONO'}
    comuns = palavras1 & palavras2 & palavras_chave
    if len(comuns) >= 2:
        return True
    
    return False

def _similaridade_procedimento(proc1: str, proc2: str, cache: Dict = None) -> float:
    """Calcula similaridade entre dois procedimentos (0.0 a 1.0) com cache e sinônimos."""
    if cache is not None:
        key = (proc1, proc2)
        if key in cache:
            return cache[key]
    
    # Expande abreviações antes de comparar
    p1_expandido = _expandir_abreviacao(proc1)
    p2_expandido = _expandir_abreviacao(proc2)
    
    if not p1_expandido or not p2_expandido:
        return 0.0
    
    # Match exato após expansão
    if p1_expandido == p2_expandido:
        if cache is not None:
            cache[(proc1, proc2)] = 1.0
        return 1.0
    
    # Verifica match por palavra contida
    if _match_palavra_contida(p1_expandido, p2_expandido):
        if cache is not None:
            cache[(proc1, proc2)] = 0.95
        return 0.95
    
    # Verifica sinônimos conhecidos
    if _verificar_sinonimo(proc1, proc2):
        if cache is not None:
            cache[(proc1, proc2)] = 1.0
        return 1.0
    
    # Verifica se um contém o outro (palavras-chave)
    palavras1 = set(p1_expandido.split())
    palavras2 = set(p2_expandido.split())
    intersecao = palavras1 & palavras2
    
    if intersecao:
        # Se tem palavras em comum, calcula similaridade
        ratio = len(intersecao) / max(len(palavras1), len(palavras2))
        if ratio >= 0.4:
            score = SequenceMatcher(None, p1_expandido, p2_expandido).ratio()
            if cache is not None:
                cache[(proc1, proc2)] = score
            return score
    
    # Fallback para SequenceMatcher completo
    score = SequenceMatcher(None, p1_expandido, p2_expandido).ratio()
    if cache is not None:
        cache[(proc1, proc2)] = score
    return score

def _datas_compativeis(data1: str, data2: str, tolerancia_dias: int = 1) -> bool:
    """Verifica se duas datas são compatíveis (tolerância ±N dias)."""
    if data1 == data2:
        return True
    try:
        d1 = datetime.strptime(data1, "%d/%m/%Y")
        d2 = datetime.strptime(data2, "%d/%m/%Y")
        return abs((d1 - d2).days) <= tolerancia_dias
    except:
        return False

def _extrair_valor_numerico(valor: str) -> float:
    """Extrai valor numérico de string."""
    if not valor or str(valor).strip() in ("", "nan", "NaN"):
        return 0.0
    try:
        s = str(valor).replace("R$", "").replace(",", ".").strip()
        return float(s)
    except:
        return 0.0

def _normalizar_nome_paciente(nome: str) -> str:
    """Normaliza nome de paciente removendo acentos e caracteres especiais."""
    if not nome or str(nome).strip() in ("", "nan", "NaN"):
        return ""
    s = str(nome).upper().strip()
    # Remove acentos
    s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    s = s.replace('Â', 'A').replace('Ê', 'E').replace('Ô', 'O')
    s = s.replace('Ã', 'A').replace('Õ', 'O')
    s = s.replace('Ç', 'C')
    s = s.replace('À', 'A').replace('È', 'E').replace('Ì', 'I').replace('Ò', 'O').replace('Ù', 'U')
    s = s.replace('Ü', 'U').replace('Ö', 'O').replace('Ä', 'A')
    # Remove caracteres especiais, mantém apenas letras e espaços
    s = re.sub(r'[^A-Z\s]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _criar_indice_repasse(df_rep: pd.DataFrame) -> Dict[Tuple[str, str], List[int]]:
    """Cria índice para busca rápida no REPASSE por (Data, Paciente)."""
    indice = {}
    for idx, row in df_rep.iterrows():
        data = row.get("Data", "")
        paciente = _normalizar_nome_paciente(row.get("Paciente", ""))
        key = (data, paciente)
        if key not in indice:
            indice[key] = []
        indice[key].append(idx)
    return indice

def _criar_indice_repasse_atendimento(df_rep: pd.DataFrame) -> Dict[Tuple[str, str], List[int]]:
    """Cria índice para busca rápida no REPASSE por (Data, NrAtendimento)."""
    indice = {}
    for idx, row in df_rep.iterrows():
        data = row.get("Data", "")
        nr_atend = str(row.get("NrAtendimento", "")).strip()
        if nr_atend and nr_atend not in ("", "nan", "NaN"):
            key = (data, nr_atend)
            if key not in indice:
                indice[key] = []
            indice[key].append(idx)
    return indice

def _determinar_status_correlacao(valor_liberado: float, tem_match: bool) -> str:
    """Determina o status de correlação baseado no valor liberado."""
    if not tem_match:
        return "NAO_FATURADO_NO_REPASSE"
    
    if valor_liberado == 0:
        return "CORRELACIONADO_COM_GLOSA_TOTAL"
    elif valor_liberado > 0:
        return "CORRELACIONADO"
    else:
        return "CORRELACIONADO"

def _extrair_tokens_nome(nome: str) -> list[str]:
    """
    Retorna os tokens significativos do nome normalizado, limitado a 4.

    Remove partículas (DE, DA, DO, DOS, DAS, E) e tokens curtos (≤ 2 chars).
    Usado pelo FB1 para gerar combinações de busca por partes do nome.

    Exemplos:
        "MARIA DAS GRACAS SILVA"  → ["MARIA", "GRACAS", "SILVA"]
        "JOAO DE SOUZA PEREIRA"   → ["JOAO", "SOUZA", "PEREIRA"]
        "ANA"                     → ["ANA"]
    """
    PARTICULAS = {"DE", "DA", "DO", "DOS", "DAS", "E"}
    partes = _normalizar_nome_paciente(nome).split()
    validos = [p for p in partes if p not in PARTICULAS and len(p) > 2]
    return validos[:4]  # limita a 4 tokens


def _tokens_fuzzy_em_comum(
    tokens_prod: list[str],
    nome_rep_norm: str,
    limiar_token: float = 0.82,
) -> list[tuple[str, str, float]]:
    """
    Para cada token da PRODUCAO, encontra o token mais similar no REPASSE
    usando SequenceMatcher. Retorna apenas os pares que atingem o limiar.

    Cada token do REPASSE só pode ser usado uma vez (match exclusivo),
    evitando que um único token do REPASSE satisfaça múltiplos tokens da PRODUCAO.

    Retorna lista de (token_prod, token_rep, score) ordenada por score desc.

    Exemplo:
        tokens_prod = ["MICHELE", "KAROLINNE", "FERNANDES"]
        nome_rep    = "MICHELLE KAROLINNE FERNANDES GUERREIRO"
        → [("MICHELE","MICHELLE",0.93), ("KAROLINNE","KAROLINNE",1.0), ("FERNANDES","FERNANDES",1.0)]
    """
    tokens_rep = [t for t in nome_rep_norm.split() if len(t) > 2]
    usados: set[str] = set()
    resultado: list[tuple[str, str, float]] = []

    for tp in tokens_prod:
        melhor_rep = None
        melhor_s = 0.0
        for tr in tokens_rep:
            if tr in usados:
                continue
            s = SequenceMatcher(None, tp, tr).ratio()
            if s >= limiar_token and s > melhor_s:
                melhor_s = s
                melhor_rep = tr
        if melhor_rep is not None:
            resultado.append((tp, melhor_rep, round(melhor_s, 3)))
            usados.add(melhor_rep)

    return sorted(resultado, key=lambda x: x[2], reverse=True)


def _buscar_fallback1_combinacoes_nome(
    data_prod: str,
    nome_prod: str,
    df_rep: pd.DataFrame,
    indice_repasse: Dict[Tuple[str, str], List[int]],
    proc_prod: str,
    cache_similaridade: Dict,
    limiar_similaridade: float = 0.65,
    limiar_token: float = 0.82,
    min_tokens_similares: int = 3,
) -> Tuple[Optional[int], float]:
    """
    Fallback 1 — Similaridade fuzzy por token + Data ±1 dia + Procedimento.

    Estratégia:
      Para cada candidato no REPASSE (dentro de ±1 dia), compara token a token
      o nome da PRODUCAO com o nome do REPASSE usando SequenceMatcher por token.
      Aceita o match se AMBAS as condições forem satisfeitas:
        1. >= min_tokens_similares tokens com similaridade >= limiar_token
        2. O primeiro token da PRODUCAO tem um similar no REPASSE
           (evita matches onde só sobrenomes coincidem)

    Por que fuzzy e não match exato:
      Typos reais como MICHELE/MICHELLE, VASCONCELOS/VASCONCELLOS,
      APPARECIDA/APARECIDA passam no limiar 0.82 e seriam perdidos
      com match exato de string.

    Por que primeiro token obrigatório:
      Impede que pares de sobrenomes genéricos (SILVA+SANTOS, PEREIRA+SILVA)
      matcheiem pacientes completamente diferentes.

    Por que min 3 tokens:
      Com apenas 2 tokens, combinações como MARIA+SILVA ou FABIANO+SILVA
      ainda geram falsos positivos frequentes em bases hospitalares.

    Parâmetros:
        limiar_token:        similaridade mínima por par de tokens (padrão 0.82)
        min_tokens_similares: quantidade mínima de tokens similares (padrão 3)

    Retorna (idx, score_procedimento) do melhor candidato, ou (None, 0.0).
    """
    tokens_prod = _extrair_tokens_nome(nome_prod)
    if len(tokens_prod) < min_tokens_similares:
        # Não há tokens suficientes para atingir o mínimo — não tenta
        return None, 0.0

    primeiro_token = tokens_prod[0]

    melhor_idx: Optional[int] = None
    melhor_score: float = 0.0

    datas_candidatas = [data_prod]
    try:
        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
        datas_candidatas += [
            (data_dt + timedelta(days=d)).strftime("%d/%m/%Y")
            for d in [-1, 1]
        ]
    except ValueError:
        pass

    for data_cand in datas_candidatas:
        for (data_rep, pac_rep), idxs in indice_repasse.items():
            if data_rep != data_cand:
                continue

            # Avalia tokens fuzzy contra o nome normalizado do candidato
            matches = _tokens_fuzzy_em_comum(tokens_prod, pac_rep, limiar_token)

            # Condição 1: mínimo de tokens similares
            if len(matches) < min_tokens_similares:
                continue

            # Condição 2: primeiro token da PRODUCAO deve ter similar no REPASSE
            primeiro_tem_match = any(tp == primeiro_token for tp, tr, s in matches)
            if not primeiro_tem_match:
                continue

            # Candidato passou nas duas condições — avalia procedimento
            for i in idxs:
                if df_rep.at[i, "_matched"]:
                    continue
                proc_rep = df_rep.iloc[i].get("Procedimento", "")
                score = _similaridade_procedimento(proc_prod, proc_rep, cache_similaridade)
                if score >= limiar_similaridade and score > melhor_score:
                    melhor_score = score
                    melhor_idx = i

    return melhor_idx, melhor_score


def _buscar_fallback2_paciente_proc_data_ampla(
    data_prod: str,
    paciente_norm: str,
    proc_prod: str,
    df_rep: pd.DataFrame,
    indice_repasse: Dict[Tuple[str, str], List[int]],
    cache_similaridade: Dict,
    limiar_similaridade: float = 0.65,
    tolerancia_dias: int = 7,
) -> Tuple[Optional[int], float]:
    """
    Fallback 2 — Nome exato + Procedimento ≥ limiar + Data ±7 dias.

    Usa o mesmo índice por (data, paciente), mas varre a janela de 2 a 7 dias
    (os ±1 dia já foram cobertos pela chave principal). O nome deve ser idêntico
    após normalização; o procedimento deve atingir o mesmo limiar de similaridade.

    Cenário típico resolvido:
        Procedimento realizado em 10/03, lançado no REPASSE como 14/03
        → ±1 dia não bate, mas ±7 dias captura.

    Retorna (idx, score) do melhor candidato, ou (None, 0.0).
    """
    melhor_idx: Optional[int] = None
    melhor_score: float = 0.0

    try:
        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
    except ValueError:
        return None, 0.0

    for delta in range(-tolerancia_dias, tolerancia_dias + 1):
        if abs(delta) <= 1:
            continue  # ±1 dia já coberto pela chave principal
        data_candidata = (data_dt + timedelta(days=delta)).strftime("%d/%m/%Y")
        key = (data_candidata, paciente_norm)
        if key not in indice_repasse:
            continue
        for i in indice_repasse[key]:
            if df_rep.at[i, "_matched"]:
                continue
            proc_rep = df_rep.iloc[i].get("Procedimento", "")
            score = _similaridade_procedimento(proc_prod, proc_rep, cache_similaridade)
            if score >= limiar_similaridade and score > melhor_score:
                melhor_score = score
                melhor_idx = i

    return melhor_idx, melhor_score


# =============================================================================
# TABELA TUSS — geração, carregamento e verificação pós-correlação
# =============================================================================

_TUSS_DIR = Path(__file__).parent
_TUSS_LOOKUP_PATH = _TUSS_DIR / "tuss_lookup_table.csv"
_TUSS_VALORES_PATH = _TUSS_DIR / "tuss_valores.csv"
_TUSS_XLSX_PATH    = _TUSS_DIR / "TUSS (ATUALIZADO).xlsx"


def _normalizar_chave_tuss(s: str) -> str:
    """Normaliza chave CONCATENAR: remove acentos, colapsa separadores e espaços."""
    import unicodedata
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[+/;,]+", "+", s)
    s = re.sub(r"\s*\+\s*", "+", s)
    return re.sub(r"\s+", " ", s).strip()


def _gerar_tabela_tuss(
    caminho_xlsx: Path = _TUSS_XLSX_PATH,
    caminho_saida: Path = _TUSS_LOOKUP_PATH,
) -> pd.DataFrame:
    """
    Lê TUSS (ATUALIZADO).xlsx e gera tuss_lookup_table.csv com classificação de tipo
    de cobrança. Retorna o DataFrame gerado. Chame sempre que o XLSX for atualizado.
    """
    import unicodedata

    df = pd.read_excel(caminho_xlsx, sheet_name=0)
    df.columns = ["Proc_PRODUCAO", "ProcAdic_PRODUCAO", "CONCATENAR", "CodigoTUSS", "Descricao_REPASSE"]
    # Remover linha de cabeçalho duplicado que entra como dado
    df = df[~df["Proc_PRODUCAO"].astype(str).str.strip().isin(["EXAME REALIZADO 1", "EXAME REALIZADO 2"])]

    # Mapa código base por procedimento principal (linha sem adicional)
    codigos_base: dict[str, str] = {}
    sem_adic = df[df["ProcAdic_PRODUCAO"].isna() | (df["ProcAdic_PRODUCAO"].astype(str).str.strip() == "")]
    for _, r in sem_adic.iterrows():
        proc_norm = _normalizar_chave_tuss(str(r["Proc_PRODUCAO"]))
        raw = str(r["CodigoTUSS"]).strip()
        sem_tuss = raw.lower() in ("nan", "sem correspondência exata tuss", "sem correspondência")
        if not sem_tuss:
            codigos_base[proc_norm] = raw.split(",")[0].strip()

    # Descrições corrigidas por conjunto de códigos (base de conhecimento curada)
    _CORRECOES_DESC: dict[frozenset, str] = {
        frozenset(["40201082"]): "Colonoscopia (Inclui A Retossigmoidoscopia)",
        frozenset(["40202038"]): "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
        frozenset(["40202186"]): "Dilatação De Esôfago Com Balão Pneumático",
        frozenset(["40202453"]): "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
        frozenset(["40202470"]): "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
        frozenset(["40202542"]): "Polipectomia De Cólon (Independente Do Número De Pólipos)",
        frozenset(["40202550"]): "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
        frozenset(["40202615"]): "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
        frozenset(["40202666"]): "Colonoscopia Com Biópsia E/Ou Citologia",
        frozenset(["40202712"]): "Colonoscopia Com Mucosectomia",
        frozenset(["40813320"]): "Colocação De Stent Biliar",
        frozenset(["40201082", "40202615"]): "Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori",
        frozenset(["40201074", "40202038"]): "Colangiopancreatografia Retrógrada Endoscópica Com Biópsia",
        frozenset(["40201104", "40202038"]): "Ecoendoscopia Alta Com Biópsia E/Ou Citologia",
        frozenset(["40202283", "40202038"]): "Gastrostomia Endoscópica Com Biópsia E/Ou Citologia",
        frozenset(["40202666", "40202542"]): "Colonoscopia Com Biópsia E Polipectomia",
        frozenset(["40202666", "40202712"]): "Colonoscopia Com Biópsia E Mucosectomia",
        frozenset(["40202666", "40202712", "40202542"]): "Colonoscopia Com Biópsia, Mucosectomia E Polipectomia",
        frozenset(["40202666", "40202135"]): "Colonoscopia Com Biópsia E Tatuagem",
        frozenset(["40202542", "40202712"]): "Colonoscopia Com Polipectomia E Mucosectomia",
        frozenset(["40202038", "40202550"]): "Endoscopia Alta Com Biópsia E Polipectomia",
        frozenset(["40202038", "40202470"]): "Endoscopia Alta Com Biópsia E Mucosectomia",
        frozenset(["40202615", "40202550"]): "Endoscopia Alta Com Urease E Polipectomia",
        frozenset(["40202615", "40202550", "40202453"]): "Endoscopia Alta Com Urease, Polipectomia E Ligadura",
        frozenset(["40202240", "40202038"]): "Ecoendoscopia Alta Com Punção E Biópsia",
        frozenset(["40202453", "40202291"]): "Endoscopia Alta Com Ligadura E Hemostasia",
        frozenset(["40202291", "40202550"]): "Endoscopia Alta Com Hemostasia E Polipectomia",
        frozenset(["40202470", "40202550"]): "Endoscopia Alta Com Mucosectomia E Polipectomia",
    }

    registros = []
    for _, r in df.iterrows():
        proc_raw  = str(r["Proc_PRODUCAO"]).strip()
        adic_raw  = str(r["ProcAdic_PRODUCAO"]).strip() if pd.notna(r["ProcAdic_PRODUCAO"]) else ""
        concat_raw = str(r["CONCATENAR"]).strip()
        codigos_raw = str(r["CodigoTUSS"]).strip()

        chave_norm = _normalizar_chave_tuss(concat_raw)
        proc_norm  = _normalizar_chave_tuss(proc_raw)
        sem_tuss   = codigos_raw.lower() in ("nan", "sem correspondência exata tuss", "sem correspondência")

        if sem_tuss:
            codigos_lista, tipo = [], "sem_mapeamento_tuss"
        else:
            codigos_lista = [c.strip() for c in codigos_raw.split(",") if c.strip()]
            base = codigos_base.get(proc_norm)
            adic_vazio = not adic_raw or adic_raw in ("nan",)
            if adic_vazio or (len(codigos_lista) == 1 and base and codigos_lista[0] == base):
                tipo = "unico_cod_tuss_somente_proc_principal"
            elif len(codigos_lista) > 1:
                tipo = "multiplos_cod_tuss_proced_adicional"
            else:
                tipo = "unico_cod_tuss_inclui_proc_adicional_e_principal"

        chave_frozen = frozenset(codigos_lista)
        desc = _CORRECOES_DESC.get(chave_frozen, str(r["Descricao_REPASSE"]).strip())

        registros.append({
            "chave_norm":              chave_norm,
            "Proc_PRODUCAO_raw":       proc_raw,
            "ProcAdic_PRODUCAO_raw":   adic_raw,
            "CONCATENAR_raw":          concat_raw,
            "CodigosTUSS":             ", ".join(codigos_lista),
            "QtdCodigos":              len(codigos_lista),
            "TipoCobranca":            tipo,
            "codigo_base_proc_principal": codigos_base.get(proc_norm, ""),
            "Descricao_REPASSE":       desc,
        })

    df_out = pd.DataFrame(registros)
    df_out.to_csv(caminho_saida, index=False, encoding="utf-8-sig")
    logger.info(f"tuss_lookup_table.csv gerado: {len(df_out)} linhas, {df_out['chave_norm'].nunique()} chaves únicas")
    return df_out


def _gerar_valores_tuss(
    df_correlacao: pd.DataFrame,
    caminho_saida: Path = _TUSS_VALORES_PATH,
    ano: int | None = None,
) -> pd.DataFrame:
    """
    Calcula estatísticas de valor por código TUSS a partir do DataFrame de correlação
    e atualiza (ou cria) tuss_valores.csv acumulando por ano.
    Chame após cada novo REPASSE processado para manter os valores atualizados.
    """
    tabela = _carregar_tabela_tuss()

    # Descrições oficiais por código individual (sem multi-código)
    _DESC_OFICIAL: dict[str, str] = {
        "40201074": "Colangiopancreatografia Retrógrada Endoscópica",
        "40201082": "Colonoscopia (Inclui A Retossigmoidoscopia)",
        "40201104": "Ecoendoscopia Alta Sem Punção",
        "40201120": "Endoscopia Digestiva Alta",
        "40201171": "Retossigmoidoscopia Flexível",
        "40201376": "Aplicação De Plasma De Argônio Por Endoscopia Digestiva Alta",
        "40202038": "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
        "40202135": "Colonoscopia Com Magnificação E Tatuagem",
        "40202143": "Descompressão Colônica Por Colonoscopia",
        "40202186": "Dilatação De Esôfago Com Balão Pneumático",
        "40202240": "Ecoendoscopia Alta Com Punção",
        "40202283": "Gastrostomia Endoscópica",
        "40202291": "Hemostasia Mecânica Do Esôfago, Estômago Ou Duodeno",
        "40202313": "Hemostasias De Cólon",
        "40202453": "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
        "40202470": "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
        "40202534": "Passagem De Sonda Naso-Enteral",
        "40202542": "Polipectomia De Cólon (Independente Do Número De Pólipos)",
        "40202550": "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
        "40202569": "Retirada De Corpo Estranho Do Cólon",
        "40202577": "Retirada De Corpo Estranho Do Esôfago, Estômago Ou Duodeno",
        "40202615": "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
        "40202666": "Colonoscopia Com Biópsia E/Ou Citologia",
        "40202682": "Retossigmoidoscopia Flexível Com Polipectomia",
        "40202690": "Retossigmoidoscopia Flexível Com Biópsia E/Ou Citologia",
        "40202712": "Colonoscopia Com Mucosectomia",
        "40813320": "Colocação De Stent Biliar",
    }

    # Extrair todos os códigos únicos da tabela
    todos_codigos: set[str] = set()
    for entry in tabela.values():
        for c in [x.strip() for x in str(entry.get("CodigosTUSS", "")).split(",") if x.strip()]:
            todos_codigos.add(c)

    # Inferir ano se não fornecido
    if ano is None:
        col_data = None
        for c in ["Data_REPASSE", "Data_PRODUCAO"]:
            if c in df_correlacao.columns:
                col_data = c
                break
        if col_data:
            try:
                anos = df_correlacao[col_data].dropna().astype(str).str.strip().str[-4:].astype(int)
                ano = int(anos.mode().iloc[0]) if len(anos) > 0 else datetime.now().year
            except Exception:
                ano = datetime.now().year
        else:
            ano = datetime.now().year

    # Preparar dados do REPASSE
    df_rep = df_correlacao.copy()
    df_rep["_cod"] = df_rep.get("CodigoTUSS_REPASSE", pd.Series(dtype=str)).astype(str).str.replace(".0", "", regex=False).str.strip()
    df_rep["_val"] = pd.to_numeric(df_rep.get("ValorLiberado_REPASSE", pd.Series(dtype=float)), errors="coerce")
    df_ano = df_rep[df_rep.get("Data_REPASSE", pd.Series(dtype=str)).astype(str).str.strip().str[-4:] == str(ano)]

    # Descrição real do REPASSE (mais frequente)
    desc_real = (
        df_rep[df_rep["_cod"].str.match(r"^\d+$", na=False)]
        .groupby(["_cod", df_rep.get("Procedimento_REPASSE", pd.Series(dtype=str)).rename("_proc")])
        .size().reset_index(name="freq")
        .sort_values(["_cod", "freq"], ascending=[True, False])
        .groupby("_cod").first().reset_index()[["_cod", "_proc"]]
    ) if "Procedimento_REPASSE" in df_correlacao.columns else pd.DataFrame(columns=["_cod", "_proc"])
    desc_real_map = dict(zip(desc_real["_cod"], desc_real["_proc"]))

    def _desc(cod):
        return desc_real_map.get(cod) or _DESC_OFICIAL.get(cod) or f"Código TUSS {cod}"

    novas_linhas = []
    for cod in sorted(todos_codigos):
        if not str(cod).isdigit():
            continue
        sub = df_ano[df_ano["_cod"] == cod]["_val"].dropna()
        qtd = len(sub)
        conf = ("Alta" if qtd >= 100 else "Boa" if qtd >= 20 else
                "Moderada" if qtd >= 5 else "Baixa" if qtd > 0 else "Sem dados")
        novas_linhas.append({
            "Ano":          ano,
            "CodigoTUSS":   cod,
            "Descricao":    _desc(cod),
            "Qtd_amostras": qtd,
            "Min":          round(float(sub.min()),    2) if qtd > 0 else None,
            "Media":        round(float(sub.mean()),   2) if qtd > 0 else None,
            "Mediana":      round(float(sub.median()), 2) if qtd > 0 else None,
            "Max":          round(float(sub.max()),    2) if qtd > 0 else None,
            "Confianca":    conf,
        })

    df_novo = pd.DataFrame(novas_linhas)

    # Acumular: preservar anos anteriores, sobrescrever o ano atual
    if caminho_saida.exists():
        df_existente = pd.read_csv(caminho_saida, dtype={"CodigoTUSS": str})
        df_existente = df_existente[df_existente["Ano"] != ano]
        df_final_val = pd.concat([df_existente, df_novo], ignore_index=True).sort_values(["Ano", "CodigoTUSS"])
    else:
        df_final_val = df_novo

    df_final_val.to_csv(caminho_saida, index=False, encoding="utf-8-sig")
    logger.info(f"tuss_valores.csv atualizado: ano={ano}, {len(df_novo)} códigos")
    return df_final_val


@st.cache_data(show_spinner=False)
def _carregar_tabela_tuss() -> dict:
    """
    Carrega tuss_lookup_table.csv como dict {chave_norm → entry}.
    Auto-gera o CSV a partir do XLSX se o arquivo não existir.
    """
    if not _TUSS_LOOKUP_PATH.exists():
        if _TUSS_XLSX_PATH.exists():
            logger.info("tuss_lookup_table.csv não encontrado — gerando a partir do XLSX")
            df = _gerar_tabela_tuss()
        else:
            logger.warning("Nem tuss_lookup_table.csv nem TUSS XLSX encontrados")
            return {}
    else:
        df = pd.read_csv(_TUSS_LOOKUP_PATH, dtype={"CodigosTUSS": str, "codigo_base_proc_principal": str})

    return {str(r["chave_norm"]): dict(r) for _, r in df.iterrows()}


@st.cache_data(show_spinner=False)
def _carregar_valores_tuss() -> dict:
    """
    Carrega tuss_valores.csv como dict {codigo → stats_ano_mais_recente}.
    Retorna vazio se arquivo não existir (estimativa desabilitada na UI).
    """
    if not _TUSS_VALORES_PATH.exists():
        logger.warning("tuss_valores.csv não encontrado — estimativas de valor desabilitadas")
        return {}
    df = pd.read_csv(_TUSS_VALORES_PATH, dtype={"CodigoTUSS": str})
    if df.empty:
        return {}
    ano_mais_recente = df["Ano"].max()
    df_atual = df[df["Ano"] == ano_mais_recente]
    return {str(r["CodigoTUSS"]): dict(r) for _, r in df_atual.iterrows()}


def _construir_indice_tuss_repasse(df_rep: pd.DataFrame) -> set:
    """Retorna set de (paciente_norm, data, codigo_tuss) para lookup O(1)."""
    idx: set = set()
    for _, r in df_rep.iterrows():
        pac  = _normalizar_nome_paciente(str(r.get("Paciente", "")))
        data = str(r.get("Data", "")).strip()
        cod  = str(r.get("CodigoTUSS", "")).replace(".0", "").strip()
        if pac and data and cod and cod != "nan":
            idx.add((pac, data, cod))
    return idx


def _construir_desc_por_codigo(df_rep: pd.DataFrame) -> dict:
    """
    Retorna dict {codigo_tuss: descricao_oficial} extraído das linhas reais do REPASSE.
    Usa a descrição mais frequente para cada código como fonte canônica.
    """
    if "CodigoTUSS" not in df_rep.columns or "Procedimento" not in df_rep.columns:
        return {}
    df_rep = df_rep.copy()
    df_rep["_cod"] = df_rep["CodigoTUSS"].astype(str).str.replace(".0", "", regex=False).str.strip()
    grp = (
        df_rep[df_rep["_cod"].str.match(r"^\d+$", na=False)]
        .groupby(["_cod", "Procedimento"]).size().reset_index(name="freq")
        .sort_values(["_cod", "freq"], ascending=[True, False])
        .groupby("_cod").first().reset_index()
    )
    return dict(zip(grp["_cod"], grp["Procedimento"]))


def verificar_tuss_adicionais(
    linhas_resultado: list,
    df_rep: pd.DataFrame,
    tabela_tuss: dict,
    tuss_idx: set,
) -> list:
    """
    Pós-processamento: para cada linha CORRELACIONADO com ProcedimentosAdicionais,
    verifica se os códigos TUSS esperados estão presentes no REPASSE.
    Adiciona colunas StatusTUSS, CodigosTUSS_Esperados e CodigosTUSS_Ausentes.
    """
    from datetime import timedelta

    for linha in linhas_resultado:
        status = str(linha.get("StatusCorrelacao", ""))
        if not status.startswith("CORRELACIONADO"):
            continue
        proc_adic = str(linha.get("ProcedimentosAdicionais_PRODUCAO", "")).strip()
        if not proc_adic or proc_adic in ("", "nan"):
            continue

        proc_princ = str(linha.get("Procedimento_PRODUCAO", "")).strip()
        chave = _normalizar_chave_tuss(f"{proc_princ}_{proc_adic}")
        entry = tabela_tuss.get(chave)

        if not entry:
            linha["StatusTUSS"] = "TUSS_COMBINACAO_SEM_MAPEAMENTO"
            continue

        tipo = str(entry.get("TipoCobranca", ""))
        codigos_raw = str(entry.get("CodigosTUSS", ""))
        codigos = [c.strip() for c in codigos_raw.split(",") if c.strip() and c.strip() != "nan"]
        base = str(entry.get("codigo_base_proc_principal", "")).replace(".0", "").strip()

        if tipo == "sem_mapeamento_tuss":
            linha["StatusTUSS"] = "TUSS_COMBINACAO_SEM_MAPEAMENTO"
            continue

        if tipo == "unico_cod_tuss_somente_proc_principal":
            linha["StatusTUSS"] = "TUSS_ADICIONAL_INCORPORADO_NO_PRINCIPAL"
            continue

        cod_repasse = str(linha.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()

        if tipo == "unico_cod_tuss_inclui_proc_adicional_e_principal":
            if codigos and cod_repasse == codigos[0]:
                linha["StatusTUSS"] = "TUSS_PROC_ADICIONAL_RECONHECIDO"
            else:
                linha["StatusTUSS"] = "TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES"
            linha["CodigosTUSS_Esperados"] = codigos[0] if codigos else ""
            continue

        # tipo == "multiplos_cod_tuss_proced_adicional"
        # Verificar cada código adicional (diferente do base) no índice do REPASSE
        pac  = _normalizar_nome_paciente(str(linha.get("Paciente_PRODUCAO", "")))
        data = str(linha.get("Data_PRODUCAO", "")).strip()
        try:
            data_dt = datetime.strptime(data, "%d/%m/%Y")
            datas_check = [(data_dt + timedelta(days=d)).strftime("%d/%m/%Y") for d in (0, -1, 1)]
        except ValueError:
            datas_check = [data]

        codigos_adicionais = [c for c in codigos if c != base]
        if not codigos_adicionais:
            codigos_adicionais = codigos

        encontrados, ausentes = [], []
        for cod in codigos_adicionais:
            achou = any((pac, d, cod) in tuss_idx for d in datas_check)
            (encontrados if achou else ausentes).append(cod)

        if not ausentes:
            linha["StatusTUSS"] = "TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS"
        else:
            linha["StatusTUSS"] = "TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE"

        linha["CodigosTUSS_Esperados"] = ", ".join(codigos_adicionais)
        if ausentes:
            linha["CodigosTUSS_Ausentes"] = ", ".join(ausentes)

    return linhas_resultado


def correlacionar_csv_arquivos(csv_producao: str, csv_repasse: str, limiar_similaridade: float = 0.65) -> str:
    """
    Correlaciona dois CSVs (PRODUCAO e REPASSE) usando chave composta otimizada.

    Garantias de schema:
    - TODAS as colunas canonicas de _MAP_PRODUCAO aparecem com sufixo _PRODUCAO
    - TODAS as colunas canonicas de _MAP_REPASSE aparecem com sufixo _REPASSE
    - Colunas ausentes no arquivo fisico ficam vazias — nunca omitidas

    Otimizacoes implementadas:
    - Indice hash para busca O(1) por Data+Paciente
    - Fallback por NrAtendimento quando nao encontra por nome
    - Normalizacao de nomes sem acentos e caracteres especiais
    - Cache de similaridade de procedimentos
    - Busca com tolerancia de +-1 dia

    Args:
        csv_producao: CSV padronizado da PRODUCAO
        csv_repasse: CSV padronizado do REPASSE
        limiar_similaridade: Threshold para match de procedimento (0.0-1.0)

    Returns:
        CSV correlacionado com sufixos _PRODUCAO e _REPASSE em todas as colunas
    """
    try:
        # ── Carrega DataFrames ────────────────────────────────────────────────
        df_prod = pd.read_csv(io.StringIO(csv_producao), dtype=str)
        df_rep  = pd.read_csv(io.StringIO(csv_repasse),  dtype=str)
        df_prod.fillna("", inplace=True)
        df_rep.fillna("",  inplace=True)

        logger.info(f"Correlacao: {len(df_prod)} linhas PRODUCAO, {len(df_rep)} linhas REPASSE")

        # ── Deriva schemas canonicos a partir dos mapas ───────────────────────
        # Todas as colunas canonicas de cada mapa (valores unicos, ordem estavel)
        _COLS_PROD = list(dict.fromkeys(
            list(_MAP_PRODUCAO.values()) +
            list(_MAP_PRODUCAO_2025_NOVO.values()) +
            list(_MAP_PRODUCAO_2026.values())
        ))
        # Remove Procedimento2 — coluna interna descartada antes de chegar aqui
        _COLS_PROD = [c for c in _COLS_PROD if c != "Procedimento2"]
        # AbaOrigemDados e adicionada pelo processador, nao consta nos mapas — inclui explicitamente
        if "AbaOrigemDados" not in _COLS_PROD:
            _COLS_PROD.append("AbaOrigemDados")

        _COLS_REP = list(dict.fromkeys(_MAP_REPASSE.values()))
        # AbaOrigemDados e adicionada pelo processador, nao consta no _MAP_REPASSE — inclui explicitamente
        if "AbaOrigemDados" not in _COLS_REP:
            _COLS_REP.append("AbaOrigemDados")

        # Colunas internas/de controle que nao devem aparecer no CSV final
        _EXCLUIR_PROD = {"TipoArquivo"}
        _EXCLUIR_REP  = {"TipoArquivo", "_matched"}

        # Garante que todas as colunas canonicas existam nos DataFrames
        # (arquivo fisico pode nao ter todas)
        for col in _COLS_PROD:
            if col not in df_prod.columns:
                df_prod[col] = ""
        for col in _COLS_REP:
            if col not in df_rep.columns:
                df_rep[col] = ""

        # ── Normaliza datas ───────────────────────────────────────────────────
        for df in [df_prod, df_rep]:
            if "Data" in df.columns:
                df["Data"] = df["Data"].apply(_padronizar_data)

        # ── Cria indices de busca rapida ──────────────────────────────────────
        indice_repasse      = _criar_indice_repasse(df_rep)
        indice_atendimento  = _criar_indice_repasse_atendimento(df_rep)
        logger.info(f"Indice por nome: {len(indice_repasse)} chaves | por atendimento: {len(indice_atendimento)} chaves")

        df_rep["_matched"] = False
        cache_similaridade: dict = {}

        linhas_resultado       = []
        matches_encontrados    = 0
        matches_por_atendimento = 0

        # ── Itera PRODUCAO e busca match no REPASSE ───────────────────────────
        for _, row_prod in df_prod.iterrows():
            data_prod     = row_prod.get("Data", "")
            paciente_norm = _normalizar_nome_paciente(row_prod.get("Paciente", ""))
            nr_atend_prod = str(row_prod.get("NrAtendimento", "")).strip()
            proc_prod     = row_prod.get("Procedimento", "")

            candidatos   = []
            metodo_busca = "1_NOME_COMPLETO_DATA_PROCEDIMENTO"

            # Busca exata por nome
            key_exata = (data_prod, paciente_norm)
            if key_exata in indice_repasse:
                candidatos.extend(indice_repasse[key_exata])

            # Busca com tolerancia de +-1 dia por nome
            if not candidatos:
                try:
                    data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
                    for delta in [-1, 1]:
                        key_tol = ((data_dt + timedelta(days=delta)).strftime("%d/%m/%Y"), paciente_norm)
                        if key_tol in indice_repasse:
                            candidatos.extend(indice_repasse[key_tol])
                except Exception:
                    pass

            # Fallback: busca por NrAtendimento
            if not candidatos and nr_atend_prod and nr_atend_prod not in ("", "nan", "NaN"):
                metodo_busca = "2_FALLBACK_NR-ATENDIMENTO_DATA_PROCEDIMENTO"
                key_atend = (data_prod, nr_atend_prod)
                if key_atend in indice_atendimento:
                    candidatos.extend(indice_atendimento[key_atend])
                if not candidatos:
                    try:
                        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
                        for delta in [-1, 1]:
                            key_tol = ((data_dt + timedelta(days=delta)).strftime("%d/%m/%Y"), nr_atend_prod)
                            if key_tol in indice_atendimento:
                                candidatos.extend(indice_atendimento[key_tol])
                    except Exception:
                        pass

            # Seleciona melhor match por similaridade de procedimento
            melhor_match = None
            melhor_score = 0.0
            melhor_idx   = None

            for idx_rep in candidatos:
                if df_rep.at[idx_rep, "_matched"]:
                    continue
                sim = _similaridade_procedimento(proc_prod, df_rep.iloc[idx_rep].get("Procedimento", ""), cache_similaridade)
                if sim >= limiar_similaridade and sim > melhor_score:
                    melhor_score = sim
                    melhor_match = df_rep.iloc[idx_rep]
                    melhor_idx   = idx_rep

            # ── Monta linha correlacionada ────────────────────────────────────
            proc_norm = _normalizar_procedimento(proc_prod)
            linha_corr: dict = {
                "ChaveCorrelacao": f"{paciente_norm}_{nr_atend_prod}_{data_prod}_{proc_norm}".replace(" ", "-")
            }

            # Todas as colunas canonicas da PRODUCAO com sufixo _PRODUCAO
            for col in _COLS_PROD:
                if col not in _EXCLUIR_PROD:
                    linha_corr[f"{col}_PRODUCAO"] = row_prod.get(col, "")

            # ── Fallbacks quando chave principal não encontrou match ──────────
            metodo_match = metodo_busca  # 1_NOME_COMPLETO ou 2_FALLBACK_NR-ATENDIMENTO

            if melhor_match is None:
                # Fallback 1: combinações de tokens do nome + data ±1 dia
                idx_fb1, score_fb1 = _buscar_fallback1_combinacoes_nome(
                    data_prod, row_prod.get("Paciente", ""),
                    df_rep, indice_repasse,
                    proc_prod, cache_similaridade,
                    limiar_similaridade,
                )
                if idx_fb1 is not None:
                    melhor_match = df_rep.iloc[idx_fb1]
                    melhor_idx   = idx_fb1
                    melhor_score = score_fb1
                    metodo_match = "3_FALLBACK_NOME_PARCIAL_FUZZY_DATA_FIXA"

            if melhor_match is None:
                # Fallback 2: nome exato + procedimento + data ±7 dias
                idx_fb2, score_fb2 = _buscar_fallback2_paciente_proc_data_ampla(
                    data_prod, paciente_norm, proc_prod,
                    df_rep, indice_repasse, cache_similaridade,
                    limiar_similaridade,
                )
                if idx_fb2 is not None:
                    melhor_match = df_rep.iloc[idx_fb2]
                    melhor_idx   = idx_fb2
                    melhor_score = score_fb2
                    metodo_match = "4_FALLBACK_NOME_COMPLETO_DATA-FLEXIVEL"

            # ── StatusCorrelacao e SimilaridadeProcedimento ───────────────────
            if melhor_match is not None:
                df_rep.at[melhor_idx, "_matched"] = True
                matches_encontrados += 1
                if metodo_busca == "2_FALLBACK_NR-ATENDIMENTO_DATA_PROCEDIMENTO":
                    matches_por_atendimento += 1

                valor_rep   = _extrair_valor_numerico(melhor_match.get("ValorLiberado", "0"))
                status_base = _determinar_status_correlacao(valor_rep, True)

                # Sufixo de rastreabilidade por método de correlação
                if metodo_match == "3_FALLBACK_NOME_PARCIAL_FUZZY_DATA_FIXA":
                    status = f"{status_base}_FALLBACK_1"
                elif metodo_match == "4_FALLBACK_NOME_COMPLETO_DATA-FLEXIVEL":
                    status = f"{status_base}_FALLBACK_2"
                elif metodo_match == "2_FALLBACK_NR-ATENDIMENTO_DATA_PROCEDIMENTO":
                    status = f"{status_base}_VIA_NR_ATENDIMENTO"
                else:
                    status = status_base

                # Ajuste A: sinaliza procedimentos anatomicamente divergentes para revisão humana
                # (ex: ENDOSCOPIA da PRODUCAO casou com "Colonoscopia" do REPASSE via
                # similaridade acidental de string — precisam ser conferidos manualmente)
                if _sao_anatomicamente_divergentes(proc_prod, melhor_match.get("Procedimento", "")):
                    status = f"{status}_PROCEDIMENTO_DIVERGENTE"

                linha_corr["SimilaridadeProcedimento"] = f"{melhor_score:.2f}"
                linha_corr["MetodoMatch"]              = metodo_match
            else:
                status = "NAO_FATURADO_NO_REPASSE"
                linha_corr["SimilaridadeProcedimento"] = "0.00"
                linha_corr["MetodoMatch"]              = "SEM_MATCH"

            linha_corr["StatusCorrelacao"] = status

            # Todas as colunas canonicas do REPASSE com sufixo _REPASSE
            if melhor_match is not None:
                for col in _COLS_REP:
                    if col not in _EXCLUIR_REP:
                        linha_corr[f"{col}_REPASSE"] = melhor_match.get(col, "")
            else:
                for col in _COLS_REP:
                    if col not in _EXCLUIR_REP:
                        linha_corr[f"{col}_REPASSE"] = ""

            linhas_resultado.append(linha_corr)

        logger.info(f"Matches encontrados: {matches_encontrados}/{len(df_prod)} ({matches_encontrados/len(df_prod)*100:.1f}%)")
        logger.info(f"Matches por atendimento: {matches_por_atendimento}")

        # ── Pré-computação para Fallbacks 5 e 6 ──────────────────────────────

        # Fallback 5: índice (paciente_norm, data) das linhas PRODUCAO já correlacionadas.
        # Permite detectar se um procedimento companion do REPASSE tem um principal
        # já matcheado no mesmo episódio (mesmo paciente + data ±1 dia).
        _corr_idx: Dict[Tuple[str, str], bool] = {}
        for lr in linhas_resultado:
            if str(lr.get("StatusCorrelacao", "")).startswith("CORRELACIONADO"):
                pac_lr  = _normalizar_nome_paciente(lr.get("Paciente_PRODUCAO", ""))
                data_lr = lr.get("Data_PRODUCAO", "")
                if pac_lr and data_lr:
                    _corr_idx[(pac_lr, data_lr)] = True

        # Fallback 6: data mínima da PRODUCAO com buffer de 30 dias.
        # Entradas do REPASSE anteriores a essa data são faturamentos tardios (late billing)
        # sem correspondência possível na PRODUCAO — recebem status informativo próprio.
        try:
            _datas_prod_parsed = [
                datetime.strptime(d.strip(), "%d/%m/%Y")
                for d in df_prod["Data"].dropna()
                if re.match(r"\d{2}/\d{2}/\d{4}", str(d).strip())
            ]
            _data_min_producao = min(_datas_prod_parsed) - timedelta(days=30) if _datas_prod_parsed else None
        except Exception:
            _data_min_producao = None

        # ── Linhas do REPASSE sem match → inseridas no final ──────────────────
        nao_matcheados = 0
        for idx_rep, row_rep in df_rep.iterrows():
            if row_rep["_matched"]:
                continue
            nao_matcheados += 1

            paciente_norm  = _normalizar_nome_paciente(row_rep.get("Paciente", ""))
            nr_atend_rep   = str(row_rep.get("NrAtendimento", "")).strip()
            data_rep       = row_rep.get("Data", "")
            proc_norm      = _normalizar_procedimento(row_rep.get("Procedimento", ""))

            linha_corr = {
                "ChaveCorrelacao": f"{paciente_norm}_{nr_atend_rep}_{data_rep}_{proc_norm}".replace(" ", "-")
            }

            # PRODUCAO vazia (schema completo, tudo vazio)
            for col in _COLS_PROD:
                if col not in _EXCLUIR_PROD:
                    linha_corr[f"{col}_PRODUCAO"] = ""

            # Determina status do REPASSE não matcheado (com fallbacks 5 e 6)
            status_repasse = "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO"
            linha_corr["MetodoMatch"] = ""

            # Fallback 5: procedimento companion (Urease/Helicobacter standalone) cujo
            # procedimento principal já foi correlacionado no mesmo episódio.
            # Condições: (a) proc é companion puro, (b) mesmo paciente tem CORRELACIONADO
            # na mesma data ±1 dia. Risco de falso positivo: muito baixo — requer
            # keyword específico + ausência de proc principal + match de episódio.
            if _e_procedimento_companion(row_rep.get("Procedimento", "")):
                try:
                    _data_rep_dt = datetime.strptime(data_rep, "%d/%m/%Y")
                    _datas_f5 = [data_rep] + [
                        (_data_rep_dt + timedelta(days=d)).strftime("%d/%m/%Y")
                        for d in [-1, 1]
                    ]
                except ValueError:
                    _datas_f5 = [data_rep]
                for _dc in _datas_f5:
                    if (paciente_norm, _dc) in _corr_idx:
                        status_repasse = "CORRELACIONADO_PROCEDIMENTO_ADICIONAL"
                        linha_corr["MetodoMatch"] = "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL"
                        break

            # Fallback 6: data do REPASSE anterior ao período coberto pela PRODUCAO
            # (faturamento tardio de exames de períodos anteriores ao arquivo enviado).
            if status_repasse == "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO" and _data_min_producao:
                try:
                    if datetime.strptime(data_rep, "%d/%m/%Y") < _data_min_producao:
                        status_repasse = "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO"
                except ValueError:
                    pass

            linha_corr["SimilaridadeProcedimento"] = "0.00"
            linha_corr["StatusCorrelacao"]         = status_repasse

            # REPASSE preenchido (schema completo)
            for col in _COLS_REP:
                if col not in _EXCLUIR_REP:
                    linha_corr[f"{col}_REPASSE"] = row_rep.get(col, "")

            linhas_resultado.append(linha_corr)

        logger.info(f"Linhas REPASSE nao matcheadas: {nao_matcheados}")
        logger.info(f"Cache de similaridade: {len(cache_similaridade)} entradas")

        # ── Verificação TUSS pós-correlação ───────────────────────────────────
        try:
            tabela_tuss = _carregar_tabela_tuss()
            if tabela_tuss:
                tuss_idx = _construir_indice_tuss_repasse(df_rep)
                linhas_resultado = verificar_tuss_adicionais(linhas_resultado, df_rep, tabela_tuss, tuss_idx)
                logger.info("Verificação TUSS concluída")
        except Exception as _e_tuss:
            logger.warning(f"Verificação TUSS ignorada: {_e_tuss}")

        # ── Monta DataFrame final e ordena por data ───────────────────────────
        df_final = pd.DataFrame(linhas_resultado)
        df_final.fillna("", inplace=True)

        if "Data_PRODUCAO" in df_final.columns:
            df_final["_sort_date"] = df_final.apply(
                lambda r: r["Data_PRODUCAO"] if r["Data_PRODUCAO"] else r.get("Data_REPASSE", ""),
                axis=1,
            )
            df_final.sort_values("_sort_date", inplace=True)
            df_final.drop(columns=["_sort_date"], inplace=True)

        return df_final.to_csv(index=False, encoding="utf-8")

    except Exception as e:
        logger.error(f"Erro na correlacao local: {e}", exc_info=True)
        return ""


# =============================================================================
# CONFIGURAÇÃO DO LLM
# =============================================================================

@st.cache_resource
def get_llm(provider: str, model_name: str, temperature: float, api_key: str, base_url: str = None) -> LLM:
    if not api_key:
        st.error("⚠️ API Key não configurada!")
        st.stop()

    model_name = model_name.strip()
    
    # Mapeamento de prefixos por provider
    provider_prefixes = {
        "Google Gemini": "google/",
        "OpenAI": "openai/",
        "Anthropic": "anthropic/",
        "Groq": "groq/",
        "Ollama": "ollama/",
        "Azure OpenAI": "azure/",
        "AWS Bedrock": "bedrock/",
        "Cohere": "cohere/",
        "HuggingFace": "huggingface/",
        "OpenRouter": "openrouter/",
        "Mistral AI": "mistral/",
        "Grok (X.AI)": "xai/",
    }
    
    prefix = provider_prefixes.get(provider, "")
    
    # Adiciona prefixo se não existir
    if prefix and not model_name.startswith(prefix):
        model_name = f"{prefix}{model_name}"
    
    try:
        llm_params = {
            "model": model_name,
            "api_key": api_key,
            "temperature": temperature
        }
        
        # Adiciona base_url se fornecido (para Ollama, Azure, etc)
        if base_url and base_url.strip():
            llm_params["base_url"] = base_url.strip()
        
        return LLM(**llm_params)
    except Exception as e:
        st.error(f"Erro ao inicializar o modelo '{model_name}': {e}")
        st.info(f"Verifique se o nome do modelo e a API Key estão corretos para {provider}")
        st.stop()


# =============================================================================
# CRIAÇÃO DOS AGENTES  ← usam session_state quando disponível
# =============================================================================

def create_agents(llm, verbose_mode: bool = False) -> list:
    cfg = _get_analista_cfg()
    multiendoscopia_analista = Agent(
        role=cfg["role"],
        goal=cfg["goal"],
        backstory=cfg["backstory"],
        llm=llm,
        verbose=verbose_mode,
        allow_delegation=False,
    )
    return [multiendoscopia_analista]


def create_correlator_agent(llm, verbose_mode: bool = False) -> Agent:
    cfg = _get_correlacionador_cfg()
    return Agent(
        role=cfg["role"],
        goal=cfg["goal"],
        backstory=cfg["backstory"],
        llm=llm,
        verbose=verbose_mode,
        allow_delegation=False,
    )


def create_correlation_task(correlator_agent: Agent, csvs_por_arquivo: dict) -> Task:
    cfg = _get_correlacionador_cfg()
    blocos = "\n\n".join(
        f"=== ARQUIVO: {fname} ===\n{csv_text}"
        for fname, csv_text in csvs_por_arquivo.items()
    )
    description = cfg["task_description_template"].replace("{blocos}", blocos)
    return Task(
        description=description,
        expected_output=cfg["task_expected_output"],
        agent=correlator_agent,
    )


# =============================================================================
# UTILITÁRIOS DE CSV
# =============================================================================

COLUNAS_ESPERADAS = [
    "Data", "NrAtendProducao", "NrAtendRepasse", "Paciente", "Convenio",
    "Procedimento", "CodigoTUSS", "MedicoExecutor",
    "QtProcedimento", "ValorLiberado", "StatusCorrelacao", "Observacao",
]


def extrair_csv_do_texto(texto: str) -> str:
    match = re.search(r"```(?:csv)?\s*\n(.*?)```", texto, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    linhas_csv = [
        linha for linha in texto.splitlines()
        if linha.count(",") >= 4 or linha.startswith("Data,")
    ]
    return "\n".join(linhas_csv).strip()


def separar_resumo_do_csv(texto_csv: str) -> tuple[str, str]:
    marcador = "# RESUMO_DIVERGENCIAS_POR_CONVENIO"
    if marcador in texto_csv:
        partes = texto_csv.split(marcador, 1)
        return partes[0].strip(), partes[1].strip()
    return texto_csv.strip(), ""


def texto_para_dataframe(csv_texto: str) -> pd.DataFrame | None:
    try:
        csv_limpo = extrair_csv_do_texto(csv_texto)
        csv_dados, _ = separar_resumo_do_csv(csv_limpo)
        df = pd.read_csv(io.StringIO(csv_dados), sep=",", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        return df
    except Exception as e:
        logger.warning(f"Falha ao converter CSV para DataFrame: {e}")
        return None


# =============================================================================
# CRIAÇÃO DAS TASKS  ← usa session_state quando disponível
# =============================================================================

def create_tasks(agents: list, conteudo_arquivo: str) -> list:
    cfg = _get_analista_cfg()
    description = cfg["task_description_template"].replace("{conteudo_arquivo}", conteudo_arquivo)
    task_analise = Task(
        description=description,
        expected_output=cfg["task_expected_output"],
        agent=agents[0],
    )
    return [task_analise]


# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================

def render_agent_form(agent_key: str, defaults: dict, title: str, icon: str):
    """
    Renderiza o formulário de edição de um agente.
    Salva os valores em st.session_state[agent_key].
    """
    cfg = st.session_state.get(agent_key, dict(defaults))

    st.markdown(f"### {icon} {title}")

    with st.container(border=True):
        col_info, col_reset = st.columns([5, 1])
        with col_info:
            st.caption("Edite os campos abaixo e clique em **Salvar Alterações** para aplicar.")
        with col_reset:
            if st.button("↩️ Restaurar", key=f"reset_{agent_key}", help="Volta aos valores padrão originais"):
                st.session_state[agent_key] = dict(defaults)
                st.rerun()

        new_role = st.text_input(
            "🏷️ Role (papel do agente)",
            value=cfg.get("role", ""),
            key=f"{agent_key}_role",
            help="Define o papel/função do agente dentro do crew.",
        )

        new_goal = st.text_area(
            "🎯 Goal (objetivo)",
            value=cfg.get("goal", ""),
            height=320,
            key=f"{agent_key}_goal",
            help="Descreve o objetivo principal do agente — instruções detalhadas de comportamento.",
        )

        new_backstory = st.text_area(
            "📖 Backstory (contexto/experiência)",
            value=cfg.get("backstory", ""),
            height=180,
            key=f"{agent_key}_backstory",
            help="Contexto e experiência do agente — usado para orientar o tom e o raciocínio.",
        )

        st.divider()
        st.markdown("#### 📋 Task associada")

        new_task_desc = st.text_area(
            "📝 Task Description (template)",
            value=cfg.get("task_description_template", ""),
            height=280,
            key=f"{agent_key}_task_desc",
            help=(
                "Template da instrução enviada ao agente em cada execução. "
                "Use {conteudo_arquivo} (Analista) ou {blocos} (Correlacionador) "
                "como placeholder do conteúdo dinâmico."
            ),
        )

        new_task_output = st.text_area(
            "✅ Task Expected Output (saída esperada)",
            value=cfg.get("task_expected_output", ""),
            height=100,
            key=f"{agent_key}_task_output",
            help="Descreve o formato/conteúdo esperado na saída da task.",
        )

        col_save, col_status = st.columns([2, 5])
        with col_save:
            if st.button(f"💾 Salvar Alterações — {title}", key=f"save_{agent_key}", type="primary"):
                st.session_state[agent_key] = {
                    "role": new_role,
                    "goal": new_goal,
                    "backstory": new_backstory,
                    "task_description_template": new_task_desc,
                    "task_expected_output": new_task_output,
                }
                with col_status:
                    st.success(f"✅ Configuração do **{title}** salva com sucesso!")


def main():
    st.set_page_config(
        page_title="Endoscopia | Controle de Procedimentos",
        page_icon="🔬",
        layout="wide",
    )

    # Inicializa os defaults dos agentes no session_state (executado uma vez por sessão)
    _init_agent_session_state()

    st.markdown(
        '<h1 style="text-align: center;">🔬 Sistema de Controle de Procedimentos de Endoscopia</h1>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="text-align: center;">'
        'Confronto automático entre Produtividade da Equipe e Repasse do Hospital<br>'
        'Identificação de Glosas, Divergências de Valor e Procedimentos Não Faturados<br>'
        'Apoio à cobrança estruturada junto ao hospital e convênios médicos'
        '</p>',
        unsafe_allow_html=True,
    )

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.info("💡 Sistema exclusivo para controle de procedimentos endoscópicos e faturamento junto a convênios.")
        st.divider()
        st.header("⚙️ Configurações")

        # Seleção do Provider
        provider = st.selectbox(
            "🤖 Provider LLM",
            options=[
                "Google Gemini",
                "OpenAI",
                "Anthropic",
                "Groq",
                "OpenRouter",
                "HuggingFace",
                "Mistral AI",
                "Grok (X.AI)",
                "Ollama",
                "Azure OpenAI",
                "AWS Bedrock",
                "Cohere",
            ],
            index=0,
            help="Selecione o provedor do modelo de linguagem"
        )

        # API Key
        env_key_map = {
            "Google Gemini": "GOOGLE_API_KEY",
            "OpenAI": "OPENAI_API_KEY",
            "Anthropic": "ANTHROPIC_API_KEY",
            "Groq": "GROQ_API_KEY",
            "OpenRouter": "OPENROUTER_API_KEY",
            "HuggingFace": "HUGGINGFACE_API_KEY",
            "Mistral AI": "MISTRAL_API_KEY",
            "Grok (X.AI)": "XAI_API_KEY",
            "Azure OpenAI": "AZURE_API_KEY",
            "AWS Bedrock": "AWS_ACCESS_KEY_ID",
            "Cohere": "COHERE_API_KEY",
            "Ollama": "",
        }

        env_key = env_key_map.get(provider, "")
        env_api_key = os.getenv(env_key, "") if env_key else ""
        
        api_key = st.text_input(
            f"🔑 API Key ({provider})",
            value=env_api_key,
            type="password",
            help=f"Sua chave de API para {provider}" + (f" (variável: {env_key})" if env_key else ""),
        )

        if api_key or provider == "Ollama":
            st.success(f"✅ {provider} configurado")
        else:
            st.warning(f"⚠️ API Key não preenchida — modo Local disponível")

        st.divider()

        # Modelos principais por provider (baseado em docs.crewai.com)
        provider_models = {
            "Google Gemini": [
                "gemini-2.5-flash-lite",
                "gemini-2.5-flash",
                "gemini-2.5-pro",
                "gemini-1.5-pro",
                "gemini-1.5-flash",
            ],
            "OpenAI": [
                "gpt-4o",
                "gpt-4o-mini",
                "gpt-4-turbo",
                "gpt-4",
                "gpt-3.5-turbo",
                "o1-preview",
                "o1-mini",
            ],
            "Anthropic": [
                "claude-3-5-sonnet-20241022",
                "claude-3-5-haiku-20241022",
                "claude-3-opus-20240229",
                "claude-3-sonnet-20240229",
                "claude-3-haiku-20240307",
            ],
            "Groq": [
                "llama-3.3-70b-versatile",
                "llama-3.1-70b-versatile",
                "llama-3.1-8b-instant",
                "mixtral-8x7b-32768",
                "gemma2-9b-it",
            ],
            "OpenRouter": [
                "anthropic/claude-3.5-sonnet",
                "openai/gpt-4o",
                "google/gemini-2.0-flash-exp",
                "meta-llama/llama-3.3-70b-instruct",
                "deepseek/deepseek-chat",
                "qwen/qwen-2.5-72b-instruct",
                "x-ai/grok-2-1212",
            ],
            "HuggingFace": [
                "meta-llama/Meta-Llama-3.1-70B-Instruct",
                "meta-llama/Meta-Llama-3.1-8B-Instruct",
                "mistralai/Mixtral-8x7B-Instruct-v0.1",
                "microsoft/Phi-3-medium-4k-instruct",
                "Qwen/Qwen2.5-72B-Instruct",
            ],
            "Mistral AI": [
                "mistral-large-latest",
                "mistral-medium-latest",
                "mistral-small-latest",
                "open-mistral-7b",
                "open-mixtral-8x7b",
            ],
            "Grok (X.AI)": [
                "grok-2-1212",
                "grok-2-vision-1212",
                "grok-beta",
            ],
            "Ollama": [
                "llama3.2",
                "llama3.1",
                "mistral",
                "qwen2.5",
                "phi3",
                "gemma2",
            ],
            "Azure OpenAI": [
                "gpt-4o",
                "gpt-4-turbo",
                "gpt-4",
                "gpt-35-turbo",
            ],
            "AWS Bedrock": [
                "anthropic.claude-3-5-sonnet-20241022-v2:0",
                "anthropic.claude-3-sonnet-20240229-v1:0",
                "anthropic.claude-3-haiku-20240307-v1:0",
                "meta.llama3-70b-instruct-v1:0",
                "meta.llama3-8b-instruct-v1:0",
            ],
            "Cohere": [
                "command-r-plus",
                "command-r",
                "command",
                "command-light",
            ],
        }

        # Nome do modelo - selectbox com opções do provider
        available_models = provider_models.get(provider, [])
        
        if available_models:
            custom_model = st.selectbox(
                "📝 Modelo",
                options=available_models,
                index=0,
                help=f"Selecione o modelo {provider}"
            )
        else:
            custom_model = st.text_input(
                "📝 Nome do Modelo",
                value="",
                help=f"Digite o nome do modelo {provider}"
            )

        # Base URL (para Ollama, Azure, etc)
        show_base_url = provider in ["Ollama", "Azure OpenAI"]
        base_url = ""
        if show_base_url:
            default_url = "http://localhost:11434" if provider == "Ollama" else ""
            base_url = st.text_input(
                "🌐 Base URL",
                value=default_url,
                help=f"URL base da API {provider}"
            )

        temperature = st.slider(
            "🌡️ Temperatura",
            min_value=0.0,
            max_value=1.0,
            value=0.1,
            step=0.1,
            help="0 = determinístico | 1 = criativo",
        )

        verbose_mode = st.toggle(
            "🔊 Modo Verbose",
            value=False,
            help="Ativado: exibe o raciocínio detalhado dos agentes (mais tokens). "
                 "Desativado: apenas o resultado final (mais econômico).",
        )

        st.divider()
        st.header("👥 Agentes Disponíveis")

        agents_info = [
            ("🔍", "Analista de Endoscopia", "Lê e estrutura os arquivos PRODUCAO e REPASSE"),
            ("🔀", "Correlacionador", "Confronta registros e identifica divergências"),
        ]
        for icon, name, desc in agents_info:
            st.markdown(f"{icon} **{name}**")
            st.caption(desc)

        st.divider()
        st.header("📌 Tipos de Arquivo")
        st.markdown("""
- **PRODUCAO**: planilha da equipe de enfermagem com os procedimentos realizados (código TUSS)
- **REPASSE**: planilha emitida pelo hospital com valores faturados e pagos pelos convênios
        """)

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab_cobranca, tab_agents = st.tabs([
        "📄 Input",
        "🚀 Execução",
        "📊 Resultados",
        "🔀 Correlação",
        "📋 Gerar Cobrança",
        "🤖 Agentes",
    ])

    # ── TAB 1: INPUT ──────────────────────────────────────────────────────────
    with tab1:
        st.header("📂 Upload de Arquivos")
        st.markdown(
            "Envie os arquivos **PRODUCAO** (equipe de enfermagem) e **REPASSE** (hospital). "
            "Formatos aceitos: **Excel (.xlsx, .xls)**, PDF, Word (.docx), TXT e CSV."
        )

        col1, col2 = st.columns([2, 1])
        extratos_texto = []

        with col1:
            uploaded_files = st.file_uploader(
                "Upload de Arquivos de Endoscopia",
                type=["xlsx", "xls", "pdf", "docx", "txt", "csv"],
                accept_multiple_files=True,
                help="Envie os arquivos PRODUCAO e REPASSE",
            )

        with col2:
            if uploaded_files:
                st.subheader("Arquivos")
                for file in uploaded_files:
                    ext = file.name.rsplit(".", 1)[-1].lower()
                    icon = "📊" if ext in ("xlsx", "xls") else "📄"
                    st.success(f"{icon} {file.name}")
                    try:
                        file.seek(0, os.SEEK_END)
                        size = file.tell() / 1024
                        file.seek(0)
                        st.caption(f"{size:.1f} KB")
                    except Exception:
                        pass

        if uploaded_files:
            st.divider()
            with st.spinner("📖 Extraindo dados dos arquivos..."):
                for file in uploaded_files:
                    text = extract_text_from_file(file)
                    if text:
                        extratos_texto.append({"filename": file.name, "content": text})

            st.success(f"✅ {len(extratos_texto)} arquivo(s) processado(s)")
            st.session_state["extratos"] = extratos_texto

        if extratos_texto:
            st.divider()
            st.subheader("👁️ Preview dos Dados")
            for doc in extratos_texto:
                with st.expander(doc["filename"]):
                    text = doc["content"]
                    st.text_area("Dados RAW", value=text, height=400)
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Caracteres", len(text))
                    c2.metric("Palavras", len(text.split()))
                    c3.metric("Linhas", len(text.splitlines()))
                    
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    nome_sem_ext = os.path.splitext(doc["filename"])[0]
                    st.download_button(
                        label="⬇️ Baixar CSV RAW",
                        data=text.encode("utf-8-sig"),
                        file_name=f"raw_{nome_sem_ext}_{ts}.csv",
                        mime="text/csv",
                        type="primary",
                    )

    # ── TAB 2: EXECUÇÃO ───────────────────────────────────────────────────────
    with tab2:
        st.header("🚀 Executar Análise dos Arquivos")

        extratos = st.session_state.get("extratos", [])
        if not extratos:
            st.warning("Envie os arquivos na aba Input antes de executar.")

        # ── Seleção do modo de execução ───────────────────────────────────────
        st.subheader("⚙️ Modo de Processamento")

        modo_execucao = st.radio(
            "Escolha como os arquivos serão processados:",
            options=[
                "🔄 Transformação Local (sem API Key)",
                f"🤖 Agente IA — {provider} (requer API Key)",
            ],
            index=0,
            horizontal=True,
            help=(
                "**Transformação Local**: usa a função `transformar_csv_arquivo` diretamente, "
                "sem consumir tokens nem precisar de API Key. Recomendado para arquivos "
                "Excel/CSV nos formatos padrão PRODUCAO e REPASSE.\n\n"
                f"**Agente IA ({provider})**: usa o agente CrewAI com o modelo configurado "
                "na sidebar. Mais flexível para arquivos fora do padrão, mas requer API Key."
            ),
        )

        usa_llm = modo_execucao.startswith("🤖")

        # Descrição contextual do modo selecionado
        if usa_llm:
            st.info(
                f"🤖 **Modo LLM ativo** — O Agente Analista de Endoscopia ({provider}) irá interpretar "
                "cada arquivo e estruturar os dados em CSV. Requer **API Key** configurada na sidebar."
            )
            if not api_key and provider != "Ollama":
                st.warning("⚠️ Preencha a **API Key** na sidebar para usar este modo.")
        else:
            st.info(
                "🔄 **Modo Local ativo** — A função `transformar_csv_arquivo` será chamada diretamente, "
                "sem uso de IA. Mais rápido e sem custo de tokens. "
                "Ideal para arquivos Excel nos formatos padrão PRODUCAO e REPASSE."
            )

        st.divider()

        btn_label = f"🚀 Iniciar Análise com {provider}" if usa_llm else "🔄 Iniciar Transformação Local"
        btn_disabled = not extratos or (usa_llm and not api_key and provider != "Ollama")

        if st.button(btn_label, type="primary", disabled=btn_disabled):

            resultados = {}
            progress_bar = st.progress(0, text="Aguardando início...")

            # ── MODO LOCAL: transformar_csv_arquivo ───────────────────────────
            if not usa_llm:
                for i, doc in enumerate(extratos):
                    filename = doc["filename"]
                    text = doc["content"]
                    total = len(extratos)

                    progress_bar.progress(
                        i / total,
                        text=f"📂 Processando arquivo {i + 1} de {total}: **{filename}**",
                    )

                    with st.status(f"🔄 Transformando: {filename}", expanded=True) as status_box:
                        try:
                            csv_resultado = transformar_csv_arquivo(text, filename)
                            if csv_resultado:
                                resultados[filename] = csv_resultado
                                status_box.update(
                                    label=f"✅ Concluído: {filename}",
                                    state="complete",
                                    expanded=False,
                                )
                            else:
                                status_box.update(
                                    label=f"⚠️ Sem dados: {filename}",
                                    state="error",
                                    expanded=True,
                                )
                                st.warning(
                                    f"Nenhum dado extraído de **{filename}**. "
                                    "Verifique se o arquivo está no formato padrão PRODUCAO ou REPASSE."
                                )
                        except Exception as exc:
                            status_box.update(
                                label=f"❌ Erro em {filename}", state="error", expanded=True
                            )
                            st.error(f"Erro ao transformar **{filename}**: {exc}")
                            logger.error(
                                f"Erro em transformar_csv_arquivo({filename}): {exc}", exc_info=True
                            )

                    progress_bar.progress(
                        (i + 1) / total,
                        text=f"✅ {i + 1} de {total} arquivos concluídos",
                    )

                st.session_state["results"] = resultados
                # No modo local o CSV já vem puro — não precisa de extração adicional
                st.session_state["csvs_brutos"] = dict(resultados)
                st.success(
                    "🎉 Transformação concluída! Veja os resultados na aba **📊 Resultados** "
                    "ou inicie a correlação na aba **🔀 Correlação**."
                )

            # ── MODO LLM: CrewAI + LLM selecionado ───────────────────────────────
            else:
                llm = get_llm(provider, custom_model, temperature, api_key, base_url)
                agents = create_agents(llm, verbose_mode)

                for i, doc in enumerate(extratos):
                    filename = doc["filename"]
                    text = doc["content"]
                    total = len(extratos)

                    progress_bar.progress(
                        i / total,
                        text=f"📂 Processando arquivo {i + 1} de {total}: **{filename}**",
                    )

                    with st.status(
                        f"🤖 Agente trabalhando em: {filename}", expanded=True
                    ) as status_box:

                        st.markdown("##### 📡 Log de Execução em Tempo Real")
                        log_container = st.container(height=380, border=False)

                        log_queue: queue.Queue = queue.Queue(maxsize=500)
                        handler = StreamlitLogHandler(log_queue)
                        handler.setFormatter(
                            logging.Formatter("%(asctime)s | %(name)s | %(message)s",
                                              datefmt="%H:%M:%S")
                        )
                        root_logger = logging.getLogger()
                        root_logger.addHandler(handler)

                        thread_result: dict = {"value": None, "error": None}

                        def _run_crew(result_holder: dict, _text=text):
                            try:
                                tasks = create_tasks(agents, _text)
                                crew = Crew(
                                    agents=agents,
                                    tasks=tasks,
                                    process=Process.sequential,
                                    verbose=verbose_mode,
                                )
                                result_holder["value"] = str(crew.kickoff())
                            except Exception as exc:
                                result_holder["error"] = exc
                                logger.error(f"Erro durante kickoff: {exc}", exc_info=True)

                        crew_thread = threading.Thread(
                            target=_run_crew,
                            args=(thread_result,),
                            daemon=True,
                        )
                        crew_thread.start()

                        while crew_thread.is_alive():
                            drained = False
                            while not log_queue.empty():
                                line = log_queue.get_nowait()
                                render_log_line(line, log_container)
                                drained = True
                            if not drained:
                                time.sleep(0.15)

                        while not log_queue.empty():
                            render_log_line(log_queue.get_nowait(), log_container)

                        crew_thread.join()
                        root_logger.removeHandler(handler)

                        if thread_result["error"]:
                            status_box.update(
                                label=f"❌ Erro em {filename}", state="error", expanded=True
                            )
                            st.error(f"Erro: {thread_result['error']}")
                        else:
                            resultados[filename] = thread_result["value"]
                            status_box.update(
                                label=f"✅ Concluído: {filename}", state="complete", expanded=False
                            )

                    progress_bar.progress(
                        (i + 1) / total,
                        text=f"✅ {i + 1} de {total} arquivos concluídos",
                    )

                st.session_state["results"] = resultados
                st.session_state["csvs_brutos"] = {
                    fname: extrair_csv_do_texto(txt)
                    for fname, txt in resultados.items()
                }
                st.success(
                    "🎉 Análise concluída! Veja os resultados na aba **📊 Resultados** "
                    "ou inicie a correlação na aba **🔀 Correlação**."
                )

    # ── TAB 3: RESULTADOS ─────────────────────────────────────────────────────
    with tab3:
        st.header("📊 Resultados por Arquivo")
        results = st.session_state.get("results", None)

        if not results:
            st.info("Execute a análise na aba Execução")
        else:
            combined_output = ""
            for file, result in results.items():
                st.subheader(file)

                st.divider()
                st.subheader("📄 Resultado Consolidado")
                st.text_area("Relatório completo", value={result}, height=400)
                c1, c2, c3 = st.columns(3)
                c1.metric("Caracteres", len(result))
                c2.metric("Palavras", len(result.split()))
                c3.metric("Linhas", len(result.splitlines()))
                
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_sem_ext = os.path.splitext(file)[0]
                st.download_button(
                    label="⬇️ Baixar CSV Processado",
                    data=result.encode("utf-8-sig"),
                    file_name=f"processado_{nome_sem_ext}_{ts}.csv",
                    mime="text/csv",
                    type="primary",
                )

    # ── TAB 4: CORRELAÇÃO ─────────────────────────────────────────────────────
    with tab4:
        st.header("🔀 Correlação — PRODUCAO × REPASSE")
        st.markdown(
            "Gera um **único CSV correlacionado** a partir de todos os arquivos processados, "
            "confrontando a produtividade da equipe com os repasses do hospital. "
            "Identifica glosas, divergências de valor e procedimentos não faturados."
        )

        csvs_brutos = st.session_state.get("csvs_brutos", {})

        if not csvs_brutos:
            st.info("⬅️ Primeiro processe os arquivos na aba **🚀 Execução**.")
        else:
            # Identifica arquivos PRODUCAO e REPASSE
            csv_producao = None
            csv_repasse = None
            nome_producao = ""
            nome_repasse = ""
            
            for fname, csv_txt in csvs_brutos.items():
                # Detecta tipo pelo conteúdo
                if "TipoArquivo" in csv_txt:
                    linhas = csv_txt.split("\n")
                    if len(linhas) > 1:
                        # Verifica primeira linha de dados
                        if "PRODUCAO" in linhas[1]:
                            csv_producao = csv_txt
                            nome_producao = fname
                        elif "REPASSE" in linhas[1]:
                            csv_repasse = csv_txt
                            nome_repasse = fname
            
            # Mostra preview dos arquivos identificados
            col_prev1, col_prev2 = st.columns(2)
            with col_prev1:
                if csv_producao:
                    st.success(f"✅ PRODUCAO: {nome_producao}")
                else:
                    st.warning("⚠️ Arquivo PRODUCAO não identificado")
            with col_prev2:
                if csv_repasse:
                    st.success(f"✅ REPASSE: {nome_repasse}")
                else:
                    st.warning("⚠️ Arquivo REPASSE não identificado")
            
            if not csv_producao or not csv_repasse:
                st.error(
                    "❌ É necessário ter exatamente 1 arquivo PRODUCAO e 1 arquivo REPASSE processados. "
                    "Verifique se os arquivos foram processados corretamente na aba Execução."
                )
            else:
                # Informações dos arquivos testados
                df_prod_prev = texto_para_dataframe(csv_producao)
                df_rep_prev = texto_para_dataframe(csv_repasse)
                
                linhas_prod = len(df_prod_prev) if df_prod_prev is not None else 0
                linhas_rep = len(df_rep_prev) if df_rep_prev is not None else 0
                
                st.markdown("### 📁 Arquivos para serem correlacionados")
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    st.metric("**PRODUCAO:**", f"{linhas_prod} linhas")
                with col_info2:
                    st.metric("**REPASSE:**", f"{linhas_rep} linhas")
                
                with st.expander(f"📋 Preview dos CSVs ({len(csvs_brutos)} arquivo(s))", expanded=False):
                    st.markdown("**PRODUCAO:**")
                    if df_prod_prev is not None and not df_prod_prev.empty:
                        st.dataframe(df_prod_prev.head(5), use_container_width=True)
                    
                    st.markdown("**REPASSE:**")
                    if df_rep_prev is not None and not df_rep_prev.empty:
                        st.dataframe(df_rep_prev.head(5), use_container_width=True)

                st.divider()
                
                # Seleção do modo de correlação
                st.subheader("⚙️ Modo de Correlação")
                
                modo_correlacao = st.radio(
                    "Escolha como realizar a correlação:",
                    options=[
                        "🔄 Correlação Local (sem API Key)",
                        f"🤖 Correlação com Agente IA — {provider} (requer API Key)",
                    ],
                    index=0,
                    horizontal=True,
                    help=(
                        "**Correlação Local**: usa a função `correlacionar_csv_arquivos` com "
                        "correspondência semântica de procedimentos. Rápido e sem custo de tokens.\n\n"
                        f"**Agente IA ({provider})**: usa o agente Correlacionador com LLM para "
                        "interpretação mais flexível. Requer API Key."
                    ),
                )
                
                usa_llm_corr = modo_correlacao.startswith("🤖")
                
                if usa_llm_corr:
                    st.info(
                        f"🤖 **Modo LLM ativo** — O Agente Correlacionador ({provider}) irá realizar "
                        "o batimento dos dados. Requer **API Key** configurada na sidebar."
                    )
                    if not api_key and provider != "Ollama":
                        st.warning("⚠️ Preencha a **API Key** na sidebar para usar este modo.")
                else:
                    st.info(
                        "🔄 **Modo Local ativo** — A função `correlacionar_csv_arquivos` será chamada "
                        "diretamente com correspondência semântica de procedimentos. Rápido e sem custo."
                    )
                
                st.divider()
                
                btn_label_corr = f"🤖 Gerar Correlação com {provider}" if usa_llm_corr else "🔄 Gerar Correlação Local"
                btn_disabled_corr = usa_llm_corr and not api_key and provider != "Ollama"

                if st.button(btn_label_corr, type="primary", disabled=btn_disabled_corr):
                    
                    # ── MODO LOCAL: correlacionar_csv_arquivos ───────────────────
                    if not usa_llm_corr:
                        with st.status("🔄 Correlacionando localmente...", expanded=True) as status_local:
                            try:
                                csv_correlacionado = correlacionar_csv_arquivos(csv_producao, csv_repasse)
                                
                                if csv_correlacionado:
                                    st.session_state["csv_correlacionado"] = csv_correlacionado
                                    status_local.update(
                                        label="✅ Correlação local concluída!",
                                        state="complete",
                                        expanded=False
                                    )
                                else:
                                    status_local.update(
                                        label="❌ Erro na correlação local",
                                        state="error",
                                        expanded=True
                                    )
                                    st.error("Não foi possível gerar a correlação. Verifique os logs.")
                            except Exception as exc:
                                status_local.update(
                                    label="❌ Erro na correlação local",
                                    state="error",
                                    expanded=True
                                )
                                st.error(f"Erro: {exc}")
                                logger.error(f"Erro em correlacionar_csv_arquivos: {exc}", exc_info=True)
                    
                    # ── MODO LLM: Agente Correlacionador ──────────────────────────
                    else:
                        with st.status("🤖 Agente Correlacionador trabalhando...", expanded=True) as status_corr:
                            st.markdown("##### 📡 Log de Correlação em Tempo Real")
                            log_container_corr = st.container(height=320, border=False)

                            log_queue_corr: queue.Queue = queue.Queue(maxsize=500)
                            handler_corr = StreamlitLogHandler(log_queue_corr)
                            handler_corr.setFormatter(
                                logging.Formatter("%(asctime)s | %(name)s | %(message)s", datefmt="%H:%M:%S")
                            )
                            root_logger = logging.getLogger()
                            root_logger.addHandler(handler_corr)

                            thread_result_corr: dict = {"value": None, "error": None}

                            def _run_correlation(result_holder: dict):
                                try:
                                    llm_corr = get_llm(
                                        provider, custom_model, temperature, api_key, base_url
                                    )
                                    correlator = create_correlator_agent(llm_corr, verbose_mode)
                                    # Passa apenas PRODUCAO e REPASSE
                                    csvs_para_correlacao = {
                                        nome_producao: csv_producao,
                                        nome_repasse: csv_repasse
                                    }
                                    task_corr = create_correlation_task(correlator, csvs_para_correlacao)
                                    crew_corr = Crew(
                                        agents=[correlator],
                                        tasks=[task_corr],
                                        process=Process.sequential,
                                        verbose=verbose_mode,
                                    )
                                    result_holder["value"] = str(crew_corr.kickoff())
                                except Exception as exc:
                                    result_holder["error"] = exc
                                    logger.error(f"Erro na correlação: {exc}", exc_info=True)

                            thread_corr = threading.Thread(
                                target=_run_correlation,
                                args=(thread_result_corr,),
                                daemon=True,
                            )
                            thread_corr.start()

                            while thread_corr.is_alive():
                                drained = False
                                while not log_queue_corr.empty():
                                    render_log_line(log_queue_corr.get_nowait(), log_container_corr)
                                    drained = True
                                if not drained:
                                    time.sleep(0.15)

                            while not log_queue_corr.empty():
                                render_log_line(log_queue_corr.get_nowait(), log_container_corr)

                            thread_corr.join()
                            root_logger.removeHandler(handler_corr)

                            if thread_result_corr["error"]:
                                status_corr.update(
                                    label="❌ Erro na correlação", state="error", expanded=True
                                )
                                st.error(f"Erro: {thread_result_corr['error']}")
                            else:
                                status_corr.update(
                                    label="✅ Correlação concluída!", state="complete", expanded=False
                                )
                                st.session_state["csv_correlacionado"] = thread_result_corr["value"]

            # ── Exibe resultado correlacionado ────────────────────────────────
            csv_correlacionado_raw = st.session_state.get("csv_correlacionado")

            if csv_correlacionado_raw:
                st.divider()

                csv_limpo = extrair_csv_do_texto(csv_correlacionado_raw)
                csv_dados_str, resumo_str = separar_resumo_do_csv(csv_limpo)

                st.subheader("📊 Tabela Correlacionada")
                df_final = texto_para_dataframe(csv_dados_str)

                if df_final is not None and not df_final.empty:
                    col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                    total_linhas = len(df_final)

                    status_col = df_final.get("StatusCorrelacao", pd.Series(dtype=str))
                    # Correlacionados = qualquer status que comece com CORRELACIONADO
                    n_correlacionado  = status_col.str.upper().str.startswith("CORRELACIONADO").sum()
                    n_divergencia     = status_col.str.upper().str.contains("DIVERGENCIA_VALOR").sum()
                    n_glosa           = status_col.str.upper().str.contains("GLOSA").sum()
                    n_nao_faturado    = (status_col.str.upper() == "NAO_FATURADO_NO_REPASSE").sum()
                    n_proc_divergente = status_col.str.upper().str.contains("PROCEDIMENTO_DIVERGENTE").sum()

                    col_m1.metric("📄 Total de Linhas",         total_linhas)
                    col_m2.metric("✅ Correlacionados",          n_correlacionado)
                    col_m3.metric("⚠️ Divergências de Valor",    int(n_divergencia))
                    col_m4.metric("🚫 Glosas",                   int(n_glosa))
                    col_m5.metric("❌ Não Faturados",             int(n_nao_faturado))
                    
                    # Métricas detalhadas de correlação
                    st.markdown("---")
                    st.markdown("### 📊 Resultados da Correlação")

                    n_repasse_nao_identificado = (status_col.str.upper() == "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO").sum()
                    n_companion               = (status_col.str.upper() == "CORRELACIONADO_PROCEDIMENTO_ADICIONAL").sum()
                    n_fora_periodo            = (status_col.str.upper() == "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO").sum()

                    # Contagem por MetodoMatch
                    mm = df_final.get("MetodoMatch", pd.Series(dtype=str)).fillna("")
                    n_m1  = (mm == "1_NOME_COMPLETO_DATA_PROCEDIMENTO").sum()
                    n_m2  = mm.str.startswith("2_FALLBACK_NR-ATENDIMENTO").sum()
                    n_m3  = (mm == "3_FALLBACK_NOME_PARCIAL_FUZZY_DATA_FIXA").sum()
                    n_m4  = (mm == "4_FALLBACK_NOME_COMPLETO_DATA-FLEXIVEL").sum()
                    n_m5  = (mm == "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL").sum()
                    n_sem = (mm == "SEM_MATCH").sum()

                    def _perc(n, base):
                        return f"{n / base * 100:.1f}%" if base > 0 else "0.0%"

                    # ── Linha 1: métricas de matches por método ───────────────
                    st.markdown("#### Matches por método de correlação")
                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric(
                        "1 Nome + Data + Proc",
                        f"{n_m1}",
                        delta=_perc(n_m1, total_linhas),
                        delta_color="off",
                        help="Nome completo normalizado + Data ±1 dia + Procedimento",
                    )
                    mc2.metric(
                        "2 NrAtendimento",
                        f"{n_m2}",
                        delta=_perc(n_m2, total_linhas),
                        delta_color="off",
                        help="NrAtendimento + Data ±1 dia + Procedimento (nome não encontrado)",
                    )
                    mc3.metric(
                        "3 Nome Fuzzy",
                        f"{n_m3}",
                        delta=_perc(n_m3, total_linhas),
                        delta_color="off",
                        help="≥3 tokens do nome com similaridade fuzzy + Data ±1 dia + Procedimento",
                    )
                    mc4.metric(
                        "4 Data Flexível",
                        f"{n_m4}",
                        delta=_perc(n_m4, total_linhas),
                        delta_color="off",
                        help="Nome completo normalizado + Data ±7 dias + Procedimento",
                    )
                    mc5.metric(
                        "5 Proc. Adicional",
                        f"{n_m5}",
                        delta=_perc(n_m5, total_linhas),
                        delta_color="off",
                        help="Urease/Helicobacter companion: principal já correlacionado no mesmo episódio",
                    )

                    # ── Linha 1b: TUSS (se disponível) ───────────────────────
                    tuss_col = df_final.get("StatusTUSS", pd.Series(dtype=str)).fillna("")
                    n_tuss_downgrade = (tuss_col.str.upper() == "TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES").sum()
                    n_tuss_ausente   = (tuss_col.str.upper() == "TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE").sum()
                    if n_tuss_downgrade + n_tuss_ausente > 0:
                        st.markdown("#### Alertas TUSS — procedimentos adicionais")
                        tc1, tc2, tc3 = st.columns(3)
                        tc1.metric(
                            "🔴 Cobrado Como Simples",
                            int(n_tuss_downgrade),
                            delta=_perc(n_tuss_downgrade, total_linhas),
                            delta_color="off",
                            help="Procedimento faturado com código simples; deveria ter código combinado (biópsia, polipectomia etc.)",
                        )
                        tc2.metric(
                            "🟠 Código Adicional Ausente",
                            int(n_tuss_ausente),
                            delta=_perc(n_tuss_ausente, total_linhas),
                            delta_color="off",
                            help="Código TUSS de procedimento adicional (ex.: Urease) não encontrado no REPASSE",
                        )
                        tc3.metric(
                            "🔎 Total p/ Cobrança",
                            int(n_tuss_downgrade + n_tuss_ausente),
                            help="Abra a aba 📋 Gerar Cobrança para exportar o formulário de revisão",
                        )

                    # ── Linha 2: pendências e alertas ─────────────────────────
                    st.markdown("#### Pendências e alertas")
                    pc1, pc2, pc3, pc4, pc5 = st.columns(5)
                    pc1.metric(
                        "❌ Não Faturados",
                        f"{int(n_nao_faturado)}",
                        delta=_perc(n_nao_faturado, total_linhas),
                        delta_color="off",
                        help="Linhas da PRODUCAO sem correspondência no REPASSE após todos os fallbacks",
                    )
                    pc2.metric(
                        "⚠️ REPASSE s/ Produção",
                        f"{int(n_repasse_nao_identificado)}",
                        delta=_perc(n_repasse_nao_identificado, total_linhas),
                        delta_color="off",
                        help="Linhas do REPASSE sem correspondência na PRODUCAO",
                    )
                    pc3.metric(
                        "🔬 Proc. Divergente",
                        f"{int(n_proc_divergente)}",
                        delta=_perc(n_proc_divergente, total_linhas),
                        delta_color="off",
                        help="Correlações onde PRODUCAO e REPASSE têm procedimentos anatomicamente distintos — revisar manualmente",
                    )
                    pc4.metric(
                        "🕰️ Fora do Período",
                        f"{int(n_fora_periodo)}",
                        delta=_perc(n_fora_periodo, total_linhas),
                        delta_color="off",
                        help="REPASSE com datas anteriores ao período coberto pela PRODUCAO (faturamento tardio)",
                    )
                    pc5.metric(
                        "🔎 Sem Match",
                        f"{int(n_sem)}",
                        delta=_perc(n_sem, total_linhas),
                        delta_color="off",
                        help="Linhas onde todos os métodos falharam",
                    )

                    st.markdown("---")

                    with st.expander("🔍 Filtros", expanded=False):
                        fcol1, fcol2, fcol3 = st.columns(3)

                        convenios_disp = sorted(df_final.get("Convenio", pd.Series()).dropna().unique().tolist())
                        filtro_convenio = fcol1.multiselect(
                            "Convênio", convenios_disp, default=convenios_disp
                        )

                        status_disp = sorted(df_final.get("StatusCorrelacao", pd.Series()).dropna().unique().tolist())
                        filtro_status = fcol2.multiselect(
                            "Status Correlação", status_disp, default=status_disp
                        )

                        tuss_disp = sorted(df_final.get("CodigoTUSS", pd.Series()).dropna().unique().tolist())
                        filtro_tuss = fcol3.multiselect(
                            "Código TUSS", tuss_disp, default=tuss_disp
                        )

                    df_filtrado = df_final.copy()
                    if "Convenio" in df_filtrado.columns and filtro_convenio:
                        df_filtrado = df_filtrado[df_filtrado["Convenio"].isin(filtro_convenio)]
                    if "StatusCorrelacao" in df_filtrado.columns and filtro_status:
                        df_filtrado = df_filtrado[df_filtrado["StatusCorrelacao"].isin(filtro_status)]
                    if "CodigoTUSS" in df_filtrado.columns and filtro_tuss:
                        df_filtrado = df_filtrado[df_filtrado["CodigoTUSS"].isin(filtro_tuss)]

                    def highlight_status(row):
                        status = str(row.get("StatusCorrelacao", "")).upper()
                        tuss   = str(row.get("StatusTUSS", "")).upper()
                        # Prioridade TUSS sobre status base quando há problema de cobrança
                        if tuss == "TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES":
                            return ["background-color: #c0392b; color: #fff"] * len(row)  # vermelho — downgrade
                        elif tuss == "TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE":
                            return ["background-color: #e67e22; color: #fff"] * len(row)  # laranja escuro — ausente
                        elif tuss == "TUSS_PROC_ADICIONAL_RECONHECIDO":
                            return ["background-color: #d4edda"] * len(row)  # verde — correto
                        elif tuss == "TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS":
                            return ["background-color: #d4edda"] * len(row)  # verde — correto
                        # Status de correlação existentes
                        if "PROCEDIMENTO_DIVERGENTE" in status:
                            return ["background-color: #ffc107; color: #000"] * len(row)  # laranja — revisar
                        elif status == "CORRELACIONADO":
                            return ["background-color: #d4edda"] * len(row)  # verde
                        elif "VIA_NR_ATENDIMENTO" in status:
                            return ["background-color: #cce5ff"] * len(row)  # azul claro
                        elif "PROCEDIMENTO_ADICIONAL" in status:
                            return ["background-color: #d4edda"] * len(row)  # verde
                        elif "FALLBACK_1" in status:
                            return ["background-color: #fff3cd"] * len(row)  # amarelo
                        elif "FALLBACK_2" in status:
                            return ["background-color: #d1ecf1"] * len(row)  # azul
                        elif "DIVERGENCIA" in status:
                            return ["background-color: #fff3cd"] * len(row)
                        elif "GLOSA" in status:
                            return ["background-color: #f8d7da"] * len(row)
                        elif "NAO_FATURADO" in status:
                            return ["background-color: #f8d7da"] * len(row)
                        elif "DATA_FORA_DO_PERIODO" in status:
                            return ["background-color: #e2e3e5; color: #888"] * len(row)  # cinza
                        elif "NAO_IDENTIFICADO" in status:
                            return ["background-color: #e2e3e5"] * len(row)
                        return [""] * len(row)

                    st.dataframe(
                        df_filtrado.style.apply(highlight_status, axis=1),
                        use_container_width=True,
                        height=460,
                    )

                    csv_download = df_filtrado.to_csv(index=False, sep=",", encoding="utf-8-sig")
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(
                        label="⬇️ Baixar CSV Correlacionado",
                        data=csv_download.encode("utf-8-sig"),
                        file_name=f"correlacao_endoscopia_{ts}.csv",
                        mime="text/csv",
                        type="primary",
                    )

                else:
                    st.warning("⚠️ Não foi possível renderizar o DataFrame. Exibindo CSV bruto.")
                    st.text_area("CSV bruto", value=csv_dados_str, height=300)
                    st.download_button(
                        label="⬇️ Baixar CSV Bruto",
                        data=csv_dados_str.encode("utf-8-sig"),
                        file_name=f"correlacao_bruta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                    )

                if resumo_str:
                    st.divider()
                    st.subheader("📋 Resumo — Divergências por Convênio")
                    st.markdown(resumo_str)

    # ── TAB 5: GERAR COBRANÇA ─────────────────────────────────────────────────
    with tab_cobranca:
        st.header("📋 Gerar Formulário de Cobrança")
        st.markdown(
            "Gera o **Formulário para Solicitação de Revisão de Procedimentos Não Repassados** "
            "preenchido automaticamente com os casos identificados na correlação."
        )

        _TEMPLATE_PATH = _TUSS_DIR / "FORMULÁRIO PARA SOLICITAÇÃO DE REVISÃO DE PROCEDIMENTOS NÃO REPASSADOS.xlsx"

        csv_corr_raw = st.session_state.get("csv_correlacionado")
        if not csv_corr_raw:
            st.info("⬅️ Execute a correlação na aba **🔀 Correlação** primeiro.")
        elif not _TEMPLATE_PATH.exists():
            st.error(f"Template não encontrado: {_TEMPLATE_PATH.name}")
        else:
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font

            # Carregar dados de correlação
            _csv_limpo  = extrair_csv_do_texto(csv_corr_raw)
            _csv_dados, _ = separar_resumo_do_csv(_csv_limpo)
            df_corr = texto_para_dataframe(_csv_dados)
            if df_corr is None or df_corr.empty:
                st.error("Não foi possível carregar os dados de correlação.")
                st.stop()

            valores_tuss  = _carregar_valores_tuss()
            desc_por_cod  = _construir_desc_por_codigo(
                pd.read_csv(BytesIO(st.session_state.get("csv_repasse_raw", b"")))
                if st.session_state.get("csv_repasse_raw") else pd.DataFrame()
            )

            # ── Descricao de código: REPASSE real > tabela DESC_OFICIAL ───────
            _DESC_OFICIAL_COB = {
                "40201074": "Colangiopancreatografia Retrógrada Endoscópica",
                "40201082": "Colonoscopia (Inclui A Retossigmoidoscopia)",
                "40201120": "Endoscopia Digestiva Alta",
                "40201171": "Retossigmoidoscopia Flexível",
                "40202038": "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
                "40202283": "Gastrostomia Endoscópica",
                "40202291": "Hemostasia Mecânica Do Esôfago, Estômago Ou Duodeno",
                "40202313": "Hemostasias De Cólon",
                "40202453": "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
                "40202470": "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
                "40202534": "Passagem De Sonda Naso-Enteral",
                "40202542": "Polipectomia De Cólon (Independente Do Número De Pólipos)",
                "40202550": "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
                "40202577": "Retirada De Corpo Estranho Do Esôfago, Estômago Ou Duodeno",
                "40202615": "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
                "40202666": "Colonoscopia Com Biópsia E/Ou Citologia",
                "40202682": "Retossigmoidoscopia Flexível Com Polipectomia",
                "40202690": "Retossigmoidoscopia Flexível Com Biópsia E/Ou Citologia",
                "40202712": "Colonoscopia Com Mucosectomia",
                "40813320": "Colocação De Stent Biliar",
            }

            def _desc_cod(cod: str) -> str:
                cod = str(cod).replace(".0", "").strip()
                return (desc_por_cod.get(cod)
                        or _DESC_OFICIAL_COB.get(cod)
                        or f"Código TUSS {cod}")

            def _valor_estimado(cod_esperado: str, cod_pago: str | None, tipo_origem: str) -> tuple[float | None, str]:
                """Retorna (valor_estimado, confiança)."""
                ce = str(cod_esperado).replace(".0", "").strip()
                cp = str(cod_pago or "").replace(".0", "").strip()
                entry_esp = valores_tuss.get(ce, {})
                if tipo_origem == "downgrade" and cp:
                    entry_pago = valores_tuss.get(cp, {})
                    m_esp  = entry_esp.get("Media")
                    m_pago = entry_pago.get("Media")
                    if m_esp is not None and m_pago is not None:
                        return round(float(m_esp) - float(m_pago), 2), str(entry_esp.get("Confianca", "Sem dados"))
                elif tipo_origem in ("ausente", "nao_faturado"):
                    m = entry_esp.get("Media") or entry_esp.get("Mediana")
                    if m is not None:
                        return round(float(m), 2), str(entry_esp.get("Confianca", "Sem dados"))
                return None, "Sem dados"

            # ── Seção 1: Filtros ──────────────────────────────────────────────
            st.subheader("1. Filtros")
            fc1, fc2 = st.columns(2)
            convenios_disp = sorted(df_corr.get("Convenio_PRODUCAO", pd.Series()).dropna().unique().tolist())
            filtro_conv = fc1.multiselect("Convênio", convenios_disp, default=convenios_disp, key="cob_conv")

            fc3, fc4, fc5 = st.columns(3)
            inc_downgrade   = fc3.checkbox("🔴 Cobrado Como Simples",  value=True, key="cob_dg",
                help="TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES — convênio usou código menor, ignorando o adicional")
            inc_ausente     = fc4.checkbox("🟠 Código Adicional Ausente", value=True, key="cob_aus",
                help="TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE — código separado (ex.: Urease) não faturado")
            inc_nao_faturado = fc5.checkbox("❌ Procedimento Não Faturado", value=True, key="cob_nf",
                help="NAO_FATURADO_NO_REPASSE — procedimento inteiro ausente do repasse")

            # ── Montar itens de cobrança ──────────────────────────────────────
            itens: list[dict] = []
            status_col  = df_corr.get("StatusCorrelacao", pd.Series(dtype=str)).fillna("")
            tuss_col_df = df_corr.get("StatusTUSS", pd.Series(dtype=str)).fillna("")

            for idx_row, row in df_corr.iterrows():
                conv = str(row.get("Convenio_PRODUCAO", "")).strip()
                if filtro_conv and conv not in filtro_conv:
                    continue
                sc   = str(row.get("StatusCorrelacao", "")).upper()
                st_t = str(row.get("StatusTUSS", "")).upper()

                # Origem 1 — downgrade
                if inc_downgrade and st_t == "TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES":
                    cod_esp  = str(row.get("CodigosTUSS_Esperados", "")).strip()
                    cod_pago = str(row.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()
                    val, conf = _valor_estimado(cod_esp, cod_pago, "downgrade")
                    obs = (f"Faturado como {cod_pago} — {_desc_cod(cod_pago)}. "
                           f"Correto conforme TUSS: {cod_esp} — {_desc_cod(cod_esp)}. "
                           f"Solicita revisão e reprocessamento.")
                    itens.append({"_origem": "downgrade", "_conf": conf,
                        "DATA": row.get("Data_PRODUCAO", ""),
                        "NR_ATEND": row.get("NrAtendimento_REPASSE", ""),
                        "PACIENTE": row.get("Paciente_PRODUCAO", ""),
                        "CONVENIO": conv,
                        "PRESTADOR": row.get("MedicoExecutor_PRODUCAO", ""),
                        "CODIGO": cod_esp,
                        "PROCEDIMENTO": _desc_cod(cod_esp),
                        "FUNCAO": row.get("Funcao_REPASSE", "Cirurgiao"),
                        "OBSERVACAO": obs,
                        "VALOR": val,
                    })

                # Origem 2 — código adicional ausente
                elif inc_ausente and st_t == "TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE":
                    ausentes_str = str(row.get("CodigosTUSS_Ausentes", row.get("CodigosTUSS_Esperados", ""))).strip()
                    for cod_aus in [c.strip() for c in ausentes_str.split(",") if c.strip()]:
                        val, conf = _valor_estimado(cod_aus, None, "ausente")
                        proc_adic = str(row.get("ProcedimentosAdicionais_PRODUCAO", "")).strip()
                        nr_base   = str(row.get("NrAtendimento_REPASSE", "")).strip()
                        obs = (f"Procedimento adicional '{proc_adic}' realizado junto ao "
                               f"{row.get('Procedimento_PRODUCAO','')} não consta no repasse. "
                               f"Código TUSS: {cod_aus}. "
                               f"Nr atendimento base: {nr_base}. Solicita inclusão.")
                        itens.append({"_origem": "ausente", "_conf": conf,
                            "DATA": row.get("Data_PRODUCAO", ""),
                            "NR_ATEND": nr_base,
                            "PACIENTE": row.get("Paciente_PRODUCAO", ""),
                            "CONVENIO": conv,
                            "PRESTADOR": row.get("MedicoExecutor_PRODUCAO", ""),
                            "CODIGO": cod_aus,
                            "PROCEDIMENTO": _desc_cod(cod_aus),
                            "FUNCAO": row.get("Funcao_REPASSE", "Cirurgiao"),
                            "OBSERVACAO": obs,
                            "VALOR": val,
                        })

                # Origem 3 — procedimento não faturado
                elif inc_nao_faturado and sc == "NAO_FATURADO_NO_REPASSE":
                    tabela_tuss_local = _carregar_tabela_tuss()
                    proc_p = str(row.get("Procedimento_PRODUCAO", "")).strip()
                    chave_base = _normalizar_chave_tuss(f"{proc_p}_")
                    entry_base = tabela_tuss_local.get(chave_base, {})
                    cod_base = str(entry_base.get("CodigosTUSS", "")).split(",")[0].strip() if entry_base else ""
                    val, conf = _valor_estimado(cod_base, None, "nao_faturado") if cod_base else (None, "Sem dados")
                    obs = (f"Procedimento {proc_p} realizado em {row.get('Data_PRODUCAO','')} "
                           f"não localizado no repasse. Solicita inclusão e faturamento.")
                    itens.append({"_origem": "nao_faturado", "_conf": conf,
                        "DATA": row.get("Data_PRODUCAO", ""),
                        "NR_ATEND": row.get("NrAtendimento_PRODUCAO", ""),
                        "PACIENTE": row.get("Paciente_PRODUCAO", ""),
                        "CONVENIO": conv,
                        "PRESTADOR": row.get("MedicoExecutor_PRODUCAO", ""),
                        "CODIGO": cod_base,
                        "PROCEDIMENTO": _desc_cod(cod_base) if cod_base else proc_p,
                        "FUNCAO": "Cirurgiao",
                        "OBSERVACAO": obs,
                        "VALOR": val,
                    })

            # ── Seção 2: Prévia ───────────────────────────────────────────────
            st.subheader("2. Prévia dos itens selecionados")
            if not itens:
                st.info("Nenhum item encontrado com os filtros selecionados.")
            else:
                df_prev = pd.DataFrame(itens)
                n_baixa_conf = (df_prev["_conf"].isin(["Baixa", "Sem dados"])).sum()
                col_res1, col_res2, col_res3 = st.columns(3)
                col_res1.metric("Total de linhas", len(df_prev))
                col_res2.metric("🔴 Cobrado Como Simples", int((df_prev["_origem"] == "downgrade").sum()))
                col_res3.metric("🟠 Código Adicional Ausente + Não Faturado",
                    int((df_prev["_origem"].isin(["ausente", "nao_faturado"])).sum()))

                colunas_prev = ["DATA", "PACIENTE", "CONVENIO", "CODIGO", "PROCEDIMENTO", "VALOR", "OBSERVACAO"]
                st.dataframe(
                    df_prev[[c for c in colunas_prev if c in df_prev.columns]],
                    use_container_width=True, height=300,
                )

                # ── Seção 3: Cabeçalho do formulário ─────────────────────────
                st.subheader("3. Cabeçalho do formulário")
                hc1, hc2, hc3 = st.columns(3)
                campo_empresa = hc1.text_input("Empresa / Prestador", value="ENDOPRIME SERVICOS MEDICOS", key="cob_empresa")
                campo_medico  = hc2.text_input("Médico Responsável", key="cob_medico")
                campo_data    = hc3.date_input("Data do formulário", value=datetime.today(), key="cob_data")

                # ── Seção 4: Opções de valor ──────────────────────────────────
                st.subheader("4. Estimativa de valor")
                estimar_valor = st.checkbox(
                    "Estimar valor da cobrança na coluna VALOR",
                    value=bool(valores_tuss),
                    disabled=not bool(valores_tuss),
                    key="cob_estimar",
                    help=(
                        "Usa a média dos valores pagos pelo convênio no período disponível em tuss_valores.csv. "
                        "Para Origem 1 (downgrade): estima a DIFERENÇA entre o código esperado e o pago. "
                        "Para Origem 2 e 3: estima o valor integral do código ausente. "
                        "Células com confiança Baixa (<5 amostras) ficam destacadas em amarelo no XLSX."
                    ),
                )
                if not valores_tuss:
                    st.caption("tuss_valores.csv não encontrado — execute _gerar_valores_tuss() para habilitar.")
                elif n_baixa_conf > 0:
                    st.caption(f"⚠️ {n_baixa_conf} item(ns) com confiança Baixa ou Sem dados serão destacados em amarelo no formulário.")

                # ── Seção 5: Gerar XLSX ───────────────────────────────────────
                st.subheader("5. Gerar formulário")
                if st.button("⚙️ Gerar Formulário XLSX", type="primary", key="cob_gerar"):
                    try:
                        wb = load_workbook(_TEMPLATE_PATH)
                        ws = wb["Plan1"]

                        # Cabeçalho
                        ws["A7"] = campo_empresa
                        ws["D7"] = campo_medico
                        ws["G7"] = f"DATA: {campo_data.strftime('%d/%m/%Y')}"

                        # Limpar dados de exemplo (linha 11 em diante)
                        for row_del in ws.iter_rows(min_row=11, max_row=ws.max_row):
                            for cell in row_del:
                                cell.value = None

                        # Estilos
                        _fill_baixa_conf = PatternFill("solid", fgColor="FFF3CD")
                        _font_normal = Font(name="Arial", size=9)
                        _font_bold   = Font(name="Arial", size=9, bold=True)

                        for i, item in enumerate(itens):
                            row_num = 11 + i
                            ws.cell(row_num, 1).value = item["DATA"]
                            ws.cell(row_num, 2).value = item["NR_ATEND"]
                            ws.cell(row_num, 3).value = str(item["PACIENTE"]).title()
                            ws.cell(row_num, 4).value = item["CONVENIO"]
                            ws.cell(row_num, 5).value = item["PRESTADOR"]
                            ws.cell(row_num, 6).value = item["CODIGO"]
                            ws.cell(row_num, 7).value = item["PROCEDIMENTO"]
                            ws.cell(row_num, 8).value = item["FUNCAO"]
                            ws.cell(row_num, 9).value = item["OBSERVACAO"]

                            if estimar_valor:
                                val = item.get("VALOR")
                                conf = item.get("_conf", "Sem dados")
                                ws.cell(row_num, 10).value = val
                                if conf in ("Baixa", "Sem dados"):
                                    ws.cell(row_num, 10).fill = _fill_baixa_conf
                            # Fonte uniforme
                            for col in range(1, 11):
                                ws.cell(row_num, col).font = _font_normal

                        # Salvar em buffer
                        buf = BytesIO()
                        wb.save(buf)
                        buf.seek(0)

                        conv_slug = (filtro_conv[0] if len(filtro_conv) == 1 else "todos").replace(" ", "_")[:20]
                        ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
                        fname = f"formulario_cobranca_{conv_slug}_{ts_now}.xlsx"

                        st.success(f"Formulário gerado: {len(itens)} linhas")
                        st.download_button(
                            label="⬇️ Baixar Formulário XLSX",
                            data=buf.getvalue(),
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                        )
                    except Exception as _e_cob:
                        st.error(f"Erro ao gerar formulário: {_e_cob}")
                        logger.error(f"Erro gerar_cobranca: {_e_cob}", exc_info=True)

    # ── TAB 6: AGENTES ────────────────────────────────────────────────────────
    with tab_agents:
        st.header("🤖 Configuração dos Agentes")
        st.markdown(
            "Visualize e edite os prompts de cada agente e suas tasks. "
            "As alterações ficam ativas **durante esta sessão** e são aplicadas "
            "imediatamente nas próximas execuções.\n\n"
            "> 💡 Use o botão **↩️ Restaurar** para voltar ao texto original a qualquer momento."
        )

        # Indicador de estado atual
        analista_modificado = st.session_state.get("analista_cfg") != ANALISTA_DEFAULTS
        correlacionador_modificado = st.session_state.get("correlacionador_cfg") != CORRELACIONADOR_DEFAULTS

        col_ind1, col_ind2, col_ind3 = st.columns([2, 2, 3])
        with col_ind1:
            if analista_modificado:
                st.warning("✏️ Analista: configuração **personalizada**")
            else:
                st.success("✅ Analista: configuração **padrão**")
        with col_ind2:
            if correlacionador_modificado:
                st.warning("✏️ Correlacionador: configuração **personalizada**")
            else:
                st.success("✅ Correlacionador: configuração **padrão**")
        with col_ind3:
            if analista_modificado or correlacionador_modificado:
                if st.button("↩️ Restaurar TODOS os agentes ao padrão", type="secondary"):
                    st.session_state["analista_cfg"] = dict(ANALISTA_DEFAULTS)
                    st.session_state["correlacionador_cfg"] = dict(CORRELACIONADOR_DEFAULTS)
                    st.rerun()

        st.divider()

        agent_tab1, agent_tab2 = st.tabs([
            "🔍 Analista de Endoscopia",
            "🔀 Correlacionador",
        ])

        with agent_tab1:
            render_agent_form(
                agent_key="analista_cfg",
                defaults=ANALISTA_DEFAULTS,
                title="Analista de Endoscopia",
                icon="🔍",
            )
            st.divider()
            with st.expander("ℹ️ Sobre o placeholder `{conteudo_arquivo}`", expanded=False):
                st.markdown("""
O campo **Task Description** do Analista usa o placeholder `{conteudo_arquivo}`.

Durante a execução, ele é substituído automaticamente pelo conteúdo extraído de cada arquivo enviado na aba **📄 Input**.

**Não remova este placeholder** — sem ele, o agente não receberá os dados do arquivo.
                """)

        with agent_tab2:
            render_agent_form(
                agent_key="correlacionador_cfg",
                defaults=CORRELACIONADOR_DEFAULTS,
                title="Correlacionador",
                icon="🔀",
            )
            st.divider()
            with st.expander("ℹ️ Sobre o placeholder `{blocos}`", expanded=False):
                st.markdown("""
O campo **Task Description** do Correlacionador usa o placeholder `{blocos}`.

Durante a execução da correlação, ele é substituído automaticamente pelos CSVs gerados pelo Analista para cada arquivo processado.

**Não remova este placeholder** — sem ele, o agente não receberá os dados para o batimento.
                """)


if __name__ == "__main__":
    main()
